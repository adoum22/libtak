from rest_framework.views import APIView
from rest_framework import generics, status, viewsets
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from drf_spectacular.utils import OpenApiParameter, OpenApiTypes, extend_schema
from django.db.models import Sum, Count, F
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal

from sales.aggregates import (
    completed_returns_for_period,
    financials_for_period,
    recognized_refund_expression,
)
from sales.models import Return, Sale
from inventory.models import Product
from core.models import AppSettings
from core.permissions import IsAdminRole, CanAccessReports
from django.http import HttpResponse
from .models import ReportSettings, ReportLog
import logging

logger = logging.getLogger(__name__)
from .serializers import (
    ReportDataSerializer,
    ReportSettingsSerializer,
    ReportLogSerializer,
)
from .tasks import get_report_data

# reportlab is imported lazily inside the PDF endpoint so the rest of the
# app keeps working on lightweight environments (e.g. PythonAnywhere free
# tier) where reportlab can't fit in the disk quota.
from io import BytesIO
from xml.sax.saxutils import escape


def _safe_spreadsheet_value(value):
    """Prevent user-controlled text from becoming an Excel formula."""
    text = '' if value is None else str(value)
    if text.startswith(('=', '+', '-', '@')):
        return "'" + text
    return text

class ExportReportView(APIView):
    """Exporter un rapport.

    PDF si reportlab est installé (poste local), sinon fallback Excel
    via openpyxl (cloud PythonAnywhere — quota disque trop serré pour
    reportlab). Le client n'a rien à changer : on sert un fichier
    téléchargeable avec le bon Content-Type.
    """
    permission_classes = [IsAuthenticated, CanAccessReports]

    @extend_schema(
        parameters=[
            OpenApiParameter('type', str, enum=['daily', 'weekly', 'monthly']),
            OpenApiParameter('format', str, enum=['pdf', 'xlsx', 'excel']),
            OpenApiParameter('date', OpenApiTypes.DATE),
            OpenApiParameter('week_offset', int),
            OpenApiParameter('month', int),
            OpenApiParameter('year', int),
        ],
        responses={
            (200, 'application/pdf'): OpenApiTypes.BINARY,
            (
                200,
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            ): OpenApiTypes.BINARY,
        },
    )
    def get(self, request):
        report_type = request.query_params.get('type', 'daily')
        today = timezone.localdate()
        if report_type not in {'daily', 'weekly', 'monthly'}:
            return Response(
                {'detail': 'Type de rapport invalide.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            if report_type == 'daily':
                date_str = request.query_params.get('date')
                start_date = end_date = (
                    datetime.strptime(date_str, '%Y-%m-%d').date()
                    if date_str else today
                )
            elif report_type == 'weekly':
                week_offset = int(request.query_params.get('week_offset', 0))
                week_offset = max(-520, min(520, week_offset))
                end_date = today - timedelta(days=7 * week_offset)
                start_date = end_date - timedelta(days=6)
            else:
                month = int(request.query_params.get('month', today.month))
                year = int(request.query_params.get('year', today.year))
                start_date = today.replace(year=year, month=month, day=1)
                if month == 12:
                    end_date = start_date.replace(year=year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    end_date = start_date.replace(month=month + 1, day=1) - timedelta(days=1)
        except (TypeError, ValueError, OverflowError):
            return Response(
                {'detail': 'Période de rapport invalide.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Données
        data = get_report_data(start_date, end_date)
        currency_symbol = str(
            AppSettings.get_settings().currency_symbol or 'DH'
        ).strip() or 'DH'

        # Forcer Excel si demandé explicitement (?format=xlsx)
        fmt = (request.query_params.get('format') or '').lower()
        if fmt in ('xlsx', 'excel'):
            return self._export_excel(report_type, start_date, end_date, data)

        try:
            # Lazy import — only load reportlab when this endpoint is hit.
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
        except ImportError:
            # Pas de reportlab (ex: PythonAnywhere free tier) -> Excel.
            logger.info("reportlab indisponible, fallback export Excel")
            return self._export_excel(report_type, start_date, end_date, data)

        try:
            # Création du PDF
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

            elements = []
            styles = getSampleStyleSheet()

            # Titre
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1e40af'),
                spaceAfter=20,
                alignment=1 # Center
            )
            elements.append(Paragraph(f"Rapport {report_type.capitalize()}", title_style))

            # Sous-titre Période
            period_style = ParagraphStyle(
                'Period',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.gray,
                alignment=1,
                spaceAfter=30
            )
            period_str = f"Période du {start_date.strftime('%d/%m/%Y')} au {end_date.strftime('%d/%m/%Y')}"
            elements.append(Paragraph(period_str, period_style))

            # Résumé (Tableau stats)
            summary_data = [
                ['Ventes', "CA", "Bénéfice Net"],
                [
                    str(data['total_sales']),
                    f"{data['total_revenue']:.2f} {currency_symbol}",
                    f"{data['total_profit']:.2f} {currency_symbol}",
                ]
            ]

            summary_table = Table(summary_data, colWidths=[5*cm, 5*cm, 5*cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, 1), colors.white),
                ('TEXTCOLOR', (0, 1), (1, 1), colors.black),
                ('TEXTCOLOR', (2, 1), (2, 1), colors.green), # Profit en vert
                ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, 1), 14),
                ('TOPPADDING', (0, 1), (-1, 1), 12),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))

            # Détail des ventes
            elements.append(Paragraph("Détail des produits vendus", styles['Heading2']))
            elements.append(Spacer(1, 10))

            # En-têtes tableau produits
            table_data = [['Produit', 'Prix moyen vendu', 'Qté', 'Total', 'Marge']]

            for item in data['items_sold']:
                table_data.append([
                    escape(item['name'][:35] + ('...' if len(item['name']) > 35 else '')),
                    f"{item.get('unit_price', 0):.2f}",
                    str(item['quantity']),
                    f"{item['revenue']:.2f}",
                    f"{item['profit']:.2f}"
                ])

            # Création tableau produits
            row_count = len(table_data)
            product_table = Table(table_data, colWidths=[7*cm, 2.5*cm, 1.5*cm, 3*cm, 3*cm])

            product_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'), # Produit aligné gauche
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'), # Chiffres alignés droite
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('TEXTCOLOR', (-1, 1), (-1, -1), colors.green), # Colonne Marge en vert
            ]))

            elements.append(product_table)
            elements.append(Spacer(1, 20))

            # Section Retours (en bas, après le tableau produits)
            if data.get('returns_count', 0) > 0:
                returns_data = [
                    ['CA Brut', 'Retours', 'CA Net'],
                    [f"{data.get('gross_revenue', 0):.2f} {currency_symbol}",
                     f"-{data.get('total_returns', 0):.2f} {currency_symbol} ({data.get('returns_count', 0)})",
                     f"{data['total_revenue']:.2f} {currency_symbol}"]
                ]

                returns_table = Table(returns_data, colWidths=[5*cm, 5*cm, 5*cm])
                returns_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef2f2')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, 1), colors.white),
                    ('TEXTCOLOR', (0, 1), (0, 1), colors.black),
                    ('TEXTCOLOR', (1, 1), (1, 1), colors.red),  # Retours en rouge
                    ('TEXTCOLOR', (2, 1), (2, 1), colors.HexColor('#1e40af')),  # CA Net en bleu
                    ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 1), (-1, 1), 12),
                    ('TOPPADDING', (0, 1), (-1, 1), 8),
                    ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#fecaca')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fecaca')),
                ]))
                elements.append(returns_table)
                elements.append(Spacer(1, 20))

            # Footer
            elements.append(Spacer(1, 20))
            footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.gray, alignment=1)
            local_time = timezone.localtime(timezone.now())
            elements.append(Paragraph(f"Généré automatiquement par Librairie App le {local_time.strftime('%d/%m/%Y à %H:%M')}", footer_style))

            # Build
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="rapport_{report_type}_{start_date}.pdf"'
            response.write(pdf)

            return response

        except Exception:
            logger.exception("PDF report generation failed, fallback Excel")
            return self._export_excel(report_type, start_date, end_date, data)

    def _export_excel(self, report_type, start_date, end_date, data):
        """Export Excel (.xlsx) du même rapport — utilisé en fallback ou
        explicitement avec ?format=xlsx. Dépend uniquement d'openpyxl,
        toujours disponible (utilisé déjà par /core/backup/)."""
        try:
            currency_symbol = str(
                AppSettings.get_settings().currency_symbol or 'DH'
            ).strip() or 'DH'
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = f"Rapport {report_type}"

            blue = Font(bold=True, color='FFFFFF', size=12)
            blue_fill = PatternFill('solid', fgColor='1E40AF')
            grey_fill = PatternFill('solid', fgColor='F3F4F6')
            green = Font(bold=True, color='16A34A')
            red = Font(bold=True, color='DC2626')
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB'),
            )

            # En-tête
            ws['A1'] = f"Rapport {report_type.capitalize()}"
            ws['A1'].font = Font(bold=True, size=16, color='1E40AF')
            ws.merge_cells('A1:E1')
            ws['A1'].alignment = Alignment(horizontal='center')

            ws['A2'] = (
                f"Période du {start_date.strftime('%d/%m/%Y')} "
                f"au {end_date.strftime('%d/%m/%Y')}"
            )
            ws.merge_cells('A2:E2')
            ws['A2'].alignment = Alignment(horizontal='center')
            ws['A2'].font = Font(italic=True, color='6B7280')

            # Bloc résumé
            row = 4
            summary = [
                ('Ventes', data.get('total_sales', 0), None),
                ('CA net', f"{data.get('total_revenue', 0):.2f} {currency_symbol}", None),
                ('Marge brute (vente - achat)',
                 f"{data.get('gross_margin', 0):.2f} {currency_symbol}", None),
                ('Dépenses d\'exploitation',
                 f"{data.get('operating_expenses', 0):.2f} {currency_symbol}", red),
                ('Bénéfice net',
                 f"{data.get('total_profit', 0):.2f} {currency_symbol}", green),
            ]
            for label, value, font in summary:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=1).fill = grey_fill
                cell = ws.cell(row=row, column=2, value=value)
                if font:
                    cell.font = font
                row += 1

            if data.get('returns_count', 0) > 0:
                ws.cell(row=row, column=1,
                        value='Retours').font = Font(bold=True)
                ws.cell(row=row, column=2,
                        value=f"-{data.get('total_returns', 0):.2f} {currency_symbol} "
                              f"({data.get('returns_count', 0)})").font = red
                row += 1

            # Tableau produits
            row += 2
            headers = ['Produit', 'Code-barres', 'Prix moyen vendu',
                       'Qté', 'CA', 'Marge']
            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = blue
                c.fill = blue_fill
                c.alignment = Alignment(horizontal='center')
                c.border = border
            row += 1

            for item in data.get('items_sold', []):
                ws.cell(
                    row=row, column=1,
                    value=_safe_spreadsheet_value(item.get('name', '')),
                )
                ws.cell(
                    row=row, column=2,
                    value=_safe_spreadsheet_value(item.get('barcode', '')),
                )
                ws.cell(row=row, column=3,
                        value=float(item.get('unit_price', 0)))
                ws.cell(row=row, column=4, value=item.get('quantity', 0))
                ws.cell(row=row, column=5,
                        value=float(item.get('revenue', 0)))
                margin_cell = ws.cell(row=row, column=6,
                                      value=float(item.get('profit', 0)))
                margin_cell.font = green
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = border
                row += 1

            # Largeurs colonnes
            widths = [38, 18, 14, 8, 14, 14]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = w

            # Pied de page
            row += 2
            local_time = timezone.localtime(timezone.now())
            ws.cell(row=row, column=1,
                    value=f"Généré le {local_time.strftime('%d/%m/%Y à %H:%M')}"
                    ).font = Font(italic=True, color='6B7280', size=9)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = (
                f'attachment; filename="rapport_{report_type}_{start_date}.xlsx"'
            )
            return response
        except Exception:
            logger.exception("Excel report generation failed")
            return Response(
                {'detail': "Erreur lors de la génération du rapport."},
                status=500,
            )


class DailyReportView(APIView):
    """Rapport journalier"""
    permission_classes = [IsAuthenticated, CanAccessReports]

    @extend_schema(
        parameters=[OpenApiParameter('date', OpenApiTypes.DATE)],
        responses=ReportDataSerializer,
    )
    def get(self, request):
        date_str = request.query_params.get('date')
        try:
            date = (
                datetime.strptime(date_str, '%Y-%m-%d').date()
                if date_str else timezone.localdate()
            )
        except (TypeError, ValueError):
            return Response(
                {'detail': 'Date invalide. Format attendu: AAAA-MM-JJ.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = get_report_data(date, date)
        data['date'] = date

        return Response(data)


class WeeklyReportView(APIView):
    """Rapport hebdomadaire"""
    permission_classes = [IsAuthenticated, CanAccessReports]

    @extend_schema(
        parameters=[OpenApiParameter('week_offset', int)],
        responses=ReportDataSerializer,
    )
    def get(self, request):
        today = timezone.localdate()

        # Semaine demandée ou courante
        try:
            week_offset = int(request.query_params.get('week_offset', 0))
        except (TypeError, ValueError):
            return Response(
                {'detail': 'Décalage de semaine invalide.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        week_offset = max(-520, min(520, week_offset))
        end_date = today - timedelta(days=7 * week_offset)
        start_date = end_date - timedelta(days=6)

        data = get_report_data(start_date, end_date)
        data['period_start'] = start_date
        data['period_end'] = end_date

        return Response(data)


class MonthlyReportView(APIView):
    """Rapport mensuel"""
    permission_classes = [IsAuthenticated, CanAccessReports]

    @extend_schema(
        parameters=[
            OpenApiParameter('month', int),
            OpenApiParameter('year', int),
        ],
        responses=ReportDataSerializer,
    )
    def get(self, request):
        today = timezone.localdate()
        try:
            month = int(request.query_params.get('month', today.month))
            year = int(request.query_params.get('year', today.year))
            start_date = today.replace(year=year, month=month, day=1)
        except (TypeError, ValueError):
            return Response(
                {'detail': 'Mois ou année invalide.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Dernier jour du mois
        if month == 12:
            end_date = start_date.replace(year=year+1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = start_date.replace(month=month+1, day=1) - timedelta(days=1)

        data = get_report_data(start_date, end_date)
        data['period_start'] = start_date
        data['period_end'] = end_date
        data['month'] = month
        data['year'] = year

        return Response(data)


class StatsView(APIView):
    """Statistiques générales"""
    permission_classes = [IsAuthenticated, CanAccessReports]

    @extend_schema(
        parameters=[OpenApiParameter('days', int)],
        responses=OpenApiTypes.OBJECT,
    )
    def get(self, request):
        today = timezone.localdate()

        # Ventes du jour
        today_sales = Sale.objects.filter(created_at__date=today)
        today_revenue = financials_for_period(today, today)['net_revenue']

        # Ventes de la semaine
        week_start = today - timedelta(days=today.weekday())
        week_sales = Sale.objects.filter(created_at__date__gte=week_start)
        week_revenue = financials_for_period(week_start, today)['net_revenue']

        # Ventes du mois
        month_start = today.replace(day=1)
        month_sales = Sale.objects.filter(created_at__date__gte=month_start)
        month_revenue = financials_for_period(month_start, today)['net_revenue']

        # Top produits: reuse the report's cent-accurate product allocation.
        # A raw SUM(SaleItem.total_price_ht) would expose gross line amounts,
        # ignoring both sale-level discounts and completed refunds.
        monthly_product_report = get_report_data(month_start, today)
        top_products = [
            {
                'product__name': row['name'],
                'product__barcode': row['barcode'],
                'total_qty': row['quantity'],
                'total_revenue': row['revenue'],
                'total_cost': row['cost'],
                'total_margin': row['profit'],
            }
            for row in monthly_product_report['items_sold'][:5]
        ]

        # Produits à réapprovisionner = stock <= min_stock (inclut ruptures)
        # On expose 3 chiffres pour qu'il n'y ait plus d'ambiguïté entre
        # "stock bas" (faible mais encore là) et "rupture" (zéro):
        #   - out_of_stock_count : stock <= 0
        #   - low_stock_only_count : 1 <= stock <= min_stock
        #   - to_replenish_count : la somme (= stock <= min_stock)
        # `low_stock_count` est conservé pour rétrocompatibilité = même
        # valeur que to_replenish_count.
        replenish_qs = Product.objects.filter(
            active=True,
            stock__lte=F('min_stock'),
        )
        out_of_stock_count = replenish_qs.filter(stock__lte=0).count()
        to_replenish_count = replenish_qs.count()
        low_stock_only_count = max(0, to_replenish_count - out_of_stock_count)
        low_stock = replenish_qs.values('id', 'name', 'stock', 'min_stock')[:10]

        # Comparaison avec hier
        yesterday = today - timedelta(days=1)
        yesterday_revenue = financials_for_period(
            yesterday, yesterday,
        )['net_revenue']

        revenue_change = 0
        if yesterday_revenue > 0:
            revenue_change = ((today_revenue - yesterday_revenue) / yesterday_revenue) * 100

        # Série N derniers jours (pour AreaChart Dashboard).
        # ?days=7 (défaut) | 30 | 90, clamp 1..365.
        from django.db.models.functions import TruncDay, TruncHour, TruncWeek
        try:
            days = int(request.query_params.get('days', 7))
        except (TypeError, ValueError):
            days = 7
        days = max(1, min(365, days))

        period_start = today - timedelta(days=days - 1)
        from credit.models import CreditPayment

        if days <= 31:
            # granularité jour
            daily_qs = (
                Sale.objects.filter(created_at__date__gte=period_start)
                .exclude(payment_method=Sale.PaymentMethod.CREDIT)
                .annotate(day=TruncDay('created_at'))
                .values('day')
                .annotate(revenue=Sum('total_ttc'), count=Count('id'))
                .order_by('day')
            )
            by_day = {row['day'].date(): row for row in daily_qs}
            credit_by_day = {
                row['day'].date(): row
                for row in CreditPayment.objects.filter(
                    created_at__date__gte=period_start,
                    status=CreditPayment.PaymentStatus.ACTIVE,
                ).annotate(day=TruncDay('created_at')).values('day').annotate(
                    revenue=Sum('amount'), count=Count('id'),
                )
                if row['day']
            }
            daily_returns = (
                completed_returns_for_period(period_start, today)
                .annotate(day=TruncDay('completed_at'))
                .values('day')
                .annotate(
                    refunds=Sum(recognized_refund_expression()),
                    returns_count=Count('id'),
                )
            )
            returns_by_day = {
                row['day'].date(): row for row in daily_returns if row['day']
            }
            revenue_7d = []
            for i in range(days):
                d = period_start + timedelta(days=i)
                row = by_day.get(d)
                credit_row = credit_by_day.get(d)
                return_row = returns_by_day.get(d)
                fmt = '%a %d/%m' if days <= 7 else '%d/%m'
                revenue_7d.append({
                    'label': d.strftime(fmt),
                    'date': d.isoformat(),
                    'revenue': float(
                        ((row['revenue'] or 0) if row else 0)
                        + ((credit_row['revenue'] or 0) if credit_row else 0)
                        - ((return_row['refunds'] or 0) if return_row else 0)
                    ),
                    'count': (row['count'] if row else 0) + (
                        credit_row['count'] if credit_row else 0
                    ),
                    'returns_count': return_row['returns_count'] if return_row else 0,
                })
        else:
            # granularité semaine pour > 1 mois (sinon trop dense)
            weekly_qs = (
                Sale.objects.filter(created_at__date__gte=period_start)
                .exclude(payment_method=Sale.PaymentMethod.CREDIT)
                .annotate(w=TruncWeek('created_at'))
                .values('w')
                .annotate(revenue=Sum('total_ttc'), count=Count('id'))
                .order_by('w')
            )
            weekly_returns = (
                completed_returns_for_period(period_start, today)
                .annotate(w=TruncWeek('completed_at'))
                .values('w')
                .annotate(
                    refunds=Sum(recognized_refund_expression()),
                    returns_count=Count('id'),
                )
            )
            sales_by_week = {
                row['w'].date(): row for row in weekly_qs if row['w']
            }
            credit_by_week = {
                row['w'].date(): row
                for row in CreditPayment.objects.filter(
                    created_at__date__gte=period_start,
                    status=CreditPayment.PaymentStatus.ACTIVE,
                ).annotate(w=TruncWeek('created_at')).values('w').annotate(
                    revenue=Sum('amount'), count=Count('id'),
                )
                if row['w']
            }
            returns_by_week = {
                row['w'].date(): row for row in weekly_returns if row['w']
            }
            revenue_7d = []
            for week_start in sorted(
                set(sales_by_week) | set(credit_by_week) | set(returns_by_week)
            ):
                row = sales_by_week.get(week_start, {})
                credit_row = credit_by_week.get(week_start, {})
                return_row = returns_by_week.get(week_start, {})
                revenue_7d.append({
                    'label': week_start.strftime('S%V'),
                    'date': week_start.isoformat(),
                    'revenue': float(
                        (row.get('revenue') or 0)
                        + (credit_row.get('revenue') or 0)
                        - (return_row.get('refunds') or 0)
                    ),
                    'count': row.get('count', 0) + credit_row.get('count', 0),
                    'returns_count': return_row.get('returns_count', 0),
                })

        # Série horaire pour aujourd'hui (BarChart)
        hourly_qs = (
            today_sales.exclude(
                payment_method=Sale.PaymentMethod.CREDIT,
            ).annotate(hour=TruncHour('created_at'))
            .values('hour')
            .annotate(revenue=Sum('total_ttc'), count=Count('id'))
            .order_by('hour')
        )
        by_hour = {row['hour'].hour: row for row in hourly_qs}
        credit_by_hour = {
            row['hour'].hour: row
            for row in CreditPayment.objects.filter(
                created_at__date=today,
                status=CreditPayment.PaymentStatus.ACTIVE,
            ).annotate(hour=TruncHour('created_at')).values('hour').annotate(
                revenue=Sum('amount'), count=Count('id'),
            )
            if row['hour']
        }
        hourly_returns = (
            completed_returns_for_period(today, today)
            .annotate(hour=TruncHour('completed_at'))
            .values('hour')
            .annotate(
                refunds=Sum(recognized_refund_expression()),
                returns_count=Count('id'),
            )
        )
        returns_by_hour = {
            row['hour'].hour: row for row in hourly_returns if row['hour']
        }
        hourly_today = []
        for h in range(24):
            row = by_hour.get(h)
            credit_row = credit_by_hour.get(h)
            return_row = returns_by_hour.get(h)
            hourly_today.append({
                'label': f'{h:02d}h',
                'revenue': float(
                    ((row['revenue'] or 0) if row else 0)
                    + ((credit_row['revenue'] or 0) if credit_row else 0)
                    - ((return_row['refunds'] or 0) if return_row else 0)
                ),
                'count': (row['count'] if row else 0) + (
                    credit_row['count'] if credit_row else 0
                ),
                'returns_count': return_row['returns_count'] if return_row else 0,
            })

        return Response({
            'today': {
                'sales_count': today_sales.count(),
                'revenue': float(today_revenue),
                'revenue_change': float(revenue_change)
            },
            'week': {
                'sales_count': week_sales.count(),
                'revenue': float(week_revenue)
            },
            'month': {
                'sales_count': month_sales.count(),
                'revenue': float(month_revenue)
            },
            'top_products': top_products,
            'low_stock': list(low_stock),
            'low_stock_count': to_replenish_count,  # rétrocompat
            'to_replenish_count': to_replenish_count,
            'low_stock_only_count': low_stock_only_count,
            'out_of_stock_count': out_of_stock_count,
            'revenue_7d': revenue_7d,
            'hourly_today': hourly_today,
        })


class ReportSettingsView(generics.RetrieveUpdateAPIView):
    """Configuration des rapports automatiques"""
    serializer_class = ReportSettingsSerializer
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get_object(self):
        return ReportSettings.get_settings()


class ReportLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Historique des rapports envoyés"""
    queryset = ReportLog.objects.all()
    serializer_class = ReportLogSerializer
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get_queryset(self):
        queryset = super().get_queryset()

        report_type = self.request.query_params.get('type')
        if report_type:
            queryset = queryset.filter(report_type=report_type)

        return queryset

    @action(detail=False, methods=['post'])
    def test_email(self, request):
        """Envoyer un rapport de test SYNCHRONE.

        Bug précédent : utilisait .delay() qui suppose Celery actif.
        Sur PythonAnywhere free tier on n'a pas Celery, .delay() levait
        une exception silencieuse et l'email n'était jamais envoyé.
        On appelle maintenant la tâche directement et on renvoie le
        statut d'envoi pour debug.
        """
        from .tasks import email_config_error, send_daily_report
        from django.conf import settings as dj_settings

        try:
            report_settings = ReportSettings.get_settings()
            previous_log_id = (
                ReportLog.objects.filter(report_type='DAILY')
                .order_by('-id')
                .values_list('id', flat=True)
                .first()
                or 0
            )
            result = send_daily_report()
            log = (
                ReportLog.objects.filter(
                    report_type='DAILY',
                    id__gt=previous_log_id,
                )
                .order_by('-sent_at')
                .first()
            )
            log_payload = {
                'sent_at': log.sent_at.isoformat() if log else None,
                'recipients': log.recipients if log else None,
                'total_sales': log.total_sales if log else None,
                'total_revenue': float(log.total_revenue) if log else None,
                'success': bool(log.success) if log else False,
                'error_message': log.error_message if log else str(result),
            }
            return Response({
                'message': result,
                'success': bool(log and log.success),
                'report_settings': {
                    'daily_enabled': report_settings.daily_enabled,
                    'recipients_count': len(report_settings.get_recipients_list()),
                    'daily_time': str(report_settings.daily_time),
                },
                'last_daily_log': log_payload,
                # Backwards compatibility for older frontends.
                'last_log': log_payload,
                'smtp_config': {
                    'backend': getattr(dj_settings, 'EMAIL_BACKEND', None),
                    'host': getattr(dj_settings, 'EMAIL_HOST', None),
                    'port': getattr(dj_settings, 'EMAIL_PORT', None),
                    'user_set': bool(getattr(dj_settings, 'EMAIL_HOST_USER', None)),
                    'password_set': bool(getattr(dj_settings, 'EMAIL_HOST_PASSWORD', None)),
                    'use_tls': getattr(dj_settings, 'EMAIL_USE_TLS', None),
                    'from_email': getattr(dj_settings, 'DEFAULT_FROM_EMAIL', None),
                    'config_error': email_config_error(),
                },
            })
        except Exception as exc:
            logger.exception('test_email failed')
            return Response(
                {
                    'message': 'Erreur envoi test',
                    'error': str(exc),
                    'success': False,
                },
                status=500,
            )

    @action(detail=False, methods=['get'])
    def diagnose(self, request):
        """Diagnostique : pourquoi le rapport peut être vide ou non envoyé.

        Vérifie en un coup d'œil :
        - settings SMTP (host/port/user/password configurés ?)
        - destinataires (vide ?)
        - daily_enabled
        - ventes vues par la même requête que le rapport (preuve qu'il y a
          bien des données)
        - dernier log d'envoi
        """
        from django.conf import settings as dj_settings
        from sales.models import Sale, SaleItem
        from .tasks import email_config_error, local_datetime_bounds

        from inventory.models import Product

        rs = ReportSettings.get_settings()
        recipients = rs.get_recipients_list()
        today = timezone.localdate()

        # Mêmes filtres que get_report_data → pour reproduire ce que voit
        # le rapport.
        start_dt, end_dt = local_datetime_bounds(today, today)
        today_sales_qs = Sale.objects.filter(
            created_at__gte=start_dt,
            created_at__lte=end_dt,
        )
        today_count = today_sales_qs.count()
        today_revenue_ttc = today_sales_qs.aggregate(
            total=Sum('total_ttc'),
        )['total'] or 0
        today_items = SaleItem.objects.filter(sale__in=today_sales_qs).count()

        # Audit du stock pour expliquer les écarts dashboard vs page Stock.
        all_products = Product.objects.values('id', 'name', 'barcode', 'stock', 'min_stock', 'active')
        out_of_stock_list = [
            p for p in all_products if (p['stock'] is None) or (p['stock'] <= 0)
        ]
        replenish_list = [
            p for p in all_products
            if p['stock'] is not None and p['min_stock'] is not None
            and p['stock'] <= p['min_stock']
        ]
        null_min_stock = [p for p in all_products if p['min_stock'] is None]
        null_stock = [p for p in all_products if p['stock'] is None]
        inactive = [p for p in all_products if not p['active']]

        last_log = ReportLog.objects.filter(report_type='DAILY').order_by('-sent_at').first()

        return Response({
            'today': today.isoformat(),
            'sales_today_count': today_count,
            'sales_today_revenue_ttc': float(today_revenue_ttc),
            'sale_items_today_count': today_items,
            'stock_audit': {
                'total_products': all_products.count(),
                'inactive_products': len(inactive),
                'out_of_stock_visible': len(out_of_stock_list),
                'out_of_stock_list': list(out_of_stock_list),
                'to_replenish_count_recompute': len(replenish_list),
                'replenish_list': list(replenish_list),
                'products_with_null_min_stock': list(null_min_stock),
                'products_with_null_stock': list(null_stock),
            },
            'report_settings': {
                'daily_enabled': rs.daily_enabled,
                'recipients': recipients,
                'recipients_count': len(recipients),
                'daily_time': str(rs.daily_time),
            },
            'smtp_config': {
                'backend': getattr(dj_settings, 'EMAIL_BACKEND', None),
                'host': getattr(dj_settings, 'EMAIL_HOST', None),
                'port': getattr(dj_settings, 'EMAIL_PORT', None),
                'user_set': bool(getattr(dj_settings, 'EMAIL_HOST_USER', None)),
                'password_set': bool(getattr(dj_settings, 'EMAIL_HOST_PASSWORD', None)),
                'use_tls': getattr(dj_settings, 'EMAIL_USE_TLS', None),
                'from_email': getattr(dj_settings, 'DEFAULT_FROM_EMAIL', None),
                'config_error': email_config_error(),
            },
            'pythonanywhere_task': (
                'cd ~/libtak/backend && '
                'python manage.py send_scheduled_reports --daily-slot'
            ),
            'pythonanywhere_env_file': '~/.libtak_env',
            'local_backup_sync_task': (
                'cd ~/libtak/backend && python manage.py local_backup_sync'
            ),
            'last_daily_log': {
                'sent_at': last_log.sent_at.isoformat() if last_log else None,
                'period_start': last_log.period_start.isoformat() if last_log else None,
                'period_end': last_log.period_end.isoformat() if last_log else None,
                'total_sales': last_log.total_sales if last_log else None,
                'total_revenue': float(last_log.total_revenue) if last_log else None,
                'recipients': last_log.recipients if last_log else None,
                'success': last_log.success if last_log else None,
                'error_message': last_log.error_message if last_log else None,
            } if last_log else None,
        })
