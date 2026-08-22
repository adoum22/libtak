from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from core.permissions import (
    CanAccessPOS,
    CanManageInventory,
    CanViewInventory,
    IsAdminRole,
    IsAdminOrReadOnly,
)
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.throttling import SimpleRateThrottle
from drf_spectacular.utils import (
    OpenApiParameter,
    OpenApiResponse,
    OpenApiTypes,
    extend_schema,
)
from django_filters.rest_framework import DjangoFilterBackend
from django.core.files.base import ContentFile
from django.shortcuts import get_object_or_404
from django.db import IntegrityError, transaction
from django.db.models import DecimalField, Prefetch, Sum, F, Q
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import PurePosixPath
import csv
import hashlib
import json
import logging
import re
import unicodedata
import uuid
import zipfile

from core.models import AuditLog

logger = logging.getLogger(__name__)

from .models import (
    Category, Product, ProductCostLayer, Supplier, StockMovement,
    PurchaseOrder, PurchaseOrderItem, PurchaseReceipt, SupplierPayment,
    InventoryCount, InventoryCountItem, PriceHistory,
)
from .serializers import (
    CategorySerializer,
    ProductSerializer,
    ProductCostLayerSerializer,
    ProductCostLayerUpdateResponseSerializer,
    ProductCreateSerializer,
    ProductImportRowSerializer,
    SupplierSerializer,
    StockMovementSerializer,
    StockInSerializer,
    BulkStockInSerializer,
    PurchaseOrderSerializer,
    PurchaseOrderCreateSerializer,
    PurchaseOrderReceiveSerializer,
    SupplierPaymentCreateSerializer,
    SupplierPaymentReverseSerializer,
    SupplierPaymentSerializer,
    InventoryCountSerializer,
    InventoryCountCreateSerializer,
    InventoryCountUpdateSerializer,
    validate_inventory_count,
)


MAX_IMPORT_FILE_SIZE = 20 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
MAX_IMPORT_ERRORS = 100
MAX_ZIP_MEMBERS = 1000
MAX_ZIP_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
MAX_IMPORT_IMAGE_SIZE = 2 * 1024 * 1024
MAX_IMPORT_IMAGE_PIXELS = 25_000_000
ALLOWED_IMPORT_IMAGE_FORMATS = {'JPEG', 'PNG', 'WEBP', 'GIF'}


def _product_audit_snapshot(product):
    """Return a stable, JSON-safe snapshot for product audit entries."""
    return {
        'name': product.name,
        'barcode': product.barcode,
        'description': product.description,
        'purchase_price': str(product.purchase_price),
        'sale_price_ht': str(product.sale_price_ht),
        'tva': str(product.tva),
        'stock': product.stock,
        'min_stock': product.min_stock,
        'category_id': product.category_id,
        'supplier_id': product.supplier_id,
        'active': product.active,
    }


class ProductImportRateThrottle(SimpleRateThrottle):
    scope = 'product_import'
    rate = '20/hour'

    def get_rate(self):
        return self.rate

    def get_cache_key(self, request, view):
        if request.user and request.user.is_authenticated:
            ident = f'user-{request.user.pk}'
        else:
            ident = self.get_ident(request)
        return self.cache_format % {'scope': self.scope, 'ident': ident}


def update_product_cost_layer_for_request(
    request,
    product_id,
    layer_id_override=None,
):
    with transaction.atomic():
        product = get_object_or_404(
            Product.objects.select_for_update(), pk=product_id
        )
        layer_id = (
            layer_id_override
            or request.data.get('layer_id')
            or request.data.get('id')
        )
        layer_id_requested = layer_id not in (None, '')
        layers = list(product.cost_layers.select_for_update().filter(
            remaining_quantity__gt=0,
        ).order_by('created_at', 'id'))
        try:
            index = int(request.data.get('index', 0) or 0)
        except (TypeError, ValueError):
            return Response({'detail': 'Position de lot invalide.'}, status=400)
        layer = None

        if layer_id not in (None, ''):
            try:
                layer_id = int(layer_id)
            except (TypeError, ValueError):
                return Response(
                    {'detail': 'Identifiant de lot FIFO invalide.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                layer = product.cost_layers.select_for_update().get(
                    id=layer_id,
                    remaining_quantity__gt=0,
                )
            except ProductCostLayer.DoesNotExist:
                layer = None

        if layer is None:
            if layer_id_requested and 'index' not in request.data:
                return Response(
                    {'detail': 'Lot FIFO actif introuvable.'},
                    status=status.HTTP_404_NOT_FOUND,
                )
            if index < 0 or index >= len(layers):
                return Response({'detail': 'Lot FIFO introuvable pour cette position.'}, status=404)
            layer = layers[index]

        old_purchase_price = product.purchase_price
        old_sale_price = product.sale_price_ht
        serializer = ProductCostLayerSerializer(
            layer,
            data=request.data,
            partial=True,
            context={'request': request},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        first_active_layer = product.cost_layers.filter(
            remaining_quantity__gt=0,
        ).order_by('created_at', 'id').first()
        product_update_fields = []
        if first_active_layer and first_active_layer.id == layer.id:
            product.purchase_price = layer.unit_cost
            product_update_fields.append('purchase_price')
        requested_sale_price = serializer.validated_data.get('sale_price')
        if requested_sale_price is not None:
            # Compatibilite de l'ancien endpoint « prix du lot » : un prix de
            # vente saisi ici devient le prix courant unique de tout le produit.
            product.sale_price_ht = requested_sale_price
            product_update_fields.append('sale_price_ht')
        if product_update_fields:
            product_update_fields.append('updated_at')
            product.save(update_fields=product_update_fields)
            if (
                product.purchase_price != old_purchase_price
                or product.sale_price_ht != old_sale_price
            ):
                PriceHistory.objects.create(
                    product=product,
                    old_purchase_price=old_purchase_price,
                    new_purchase_price=product.purchase_price,
                    old_sale_price=old_sale_price,
                    new_sale_price=product.sale_price_ht,
                    changed_by=request.user,
                    reason='Correction du lot FIFO actif',
                )
        product.refresh_from_db()
    return Response({
        'layer': serializer.data,
        'product': ProductSerializer(product, context={'request': request}).data,
    })


@extend_schema(
    request=ProductCostLayerSerializer,
    responses={
        200: ProductCostLayerUpdateResponseSerializer,
        400: OpenApiResponse(description='Données de lot invalides.'),
        404: OpenApiResponse(description='Produit ou lot FIFO actif introuvable.'),
    },
)
@api_view(['PATCH', 'POST'])
@permission_classes([IsAuthenticated, IsAdminRole])
def update_product_cost_layer(request, product_id):
    """Endpoint explicite et stable pour corriger un lot FIFO produit."""
    return update_product_cost_layer_for_request(request, product_id)


class SupplierSearchFilter(filters.SearchFilter):
    """Do not let hidden contact fields become a search oracle."""

    def get_search_fields(self, view, request):
        if request.user and request.user.is_admin_role:
            return super().get_search_fields(view, request)
        return ['name']


class SupplierViewSet(viewsets.ModelViewSet):
    """API pour les fournisseurs"""
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated, CanViewInventory, IsAdminOrReadOnly]
    filter_backends = [SupplierSearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'contact_name', 'email', 'phone']
    ordering_fields = ['name', 'created_at']
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()
        active = self.request.query_params.get('active')
        if active is not None:
            queryset = queryset.filter(active=active.lower() == 'true')
        return queryset


    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            supplier = get_object_or_404(
                Supplier.objects.select_for_update(),
                pk=kwargs['pk'],
            )
            self.check_object_permissions(request, supplier)
            supplier.active = False
            supplier.save(update_fields=['active', 'updated_at'])
        return Response(status=status.HTTP_204_NO_CONTENT)


class CategoryViewSet(viewsets.ModelViewSet):
    """API pour les catégories"""
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated, CanViewInventory, IsAdminOrReadOnly]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name']


class ProductViewSet(viewsets.ModelViewSet):
    """API pour les produits"""
    queryset = Product.objects.select_related('category', 'supplier').prefetch_related('cost_layers').all()
    permission_classes = [IsAuthenticated, CanManageInventory]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'supplier', 'active']
    search_fields = ['name', 'barcode', 'description']
    ordering_fields = ['name', 'stock', 'sale_price_ht', 'created_at']
    ordering = ['name']

    def _parse_decimal_param(self, name):
        value = self.request.query_params.get(name)
        if value in (None, ''):
            return None
        try:
            return Decimal(str(value).replace(',', '.'))
        except (InvalidOperation, TypeError):
            return None

    def get_serializer_class(self):
        if self.action == 'create':
            return ProductCreateSerializer
        return ProductSerializer

    def get_queryset(self):
        queryset = super().get_queryset()

        if self.action == 'pos':
            queryset = queryset.filter(
                active=True,
                sale_price_ht__gt=0,
            )

        # Filtre par code-barres exact
        barcode = self.request.query_params.get('barcode')
        if barcode:
            queryset = queryset.filter(barcode=barcode)

        # Filtre stock bas (incluant les ruptures)
        low_stock = self.request.query_params.get('low_stock')
        if low_stock and low_stock.lower() == 'true':
            queryset = queryset.filter(stock__lte=F('min_stock'))

        # Filtre rupture seulement (stock = 0)
        stock_status = self.request.query_params.get('stock_status')
        if stock_status == 'out':
            queryset = queryset.filter(stock__lte=0)
        elif stock_status == 'low':
            queryset = queryset.filter(stock__lte=F('min_stock'), stock__gt=0)

        is_admin = self.request.user.is_admin_role
        purchase_price = (
            self._parse_decimal_param('purchase_price') if is_admin else None
        )
        purchase_price_min = (
            self._parse_decimal_param('purchase_price_min') if is_admin else None
        )
        purchase_price_max = (
            self._parse_decimal_param('purchase_price_max') if is_admin else None
        )

        # Recherche prix achat sur le prix par defaut du produit OU sur les lots
        # FIFO actifs. Utile pour retrouver les produits inventaires a 0 / 1000 DH.
        if purchase_price is not None:
            queryset = queryset.filter(
                Q(purchase_price=purchase_price)
                | Q(
                    cost_layers__remaining_quantity__gt=0,
                    cost_layers__unit_cost=purchase_price,
                )
            ).distinct()
        if purchase_price_min is not None:
            queryset = queryset.filter(
                Q(purchase_price__gte=purchase_price_min)
                | Q(
                    cost_layers__remaining_quantity__gt=0,
                    cost_layers__unit_cost__gte=purchase_price_min,
                )
            ).distinct()
        if purchase_price_max is not None:
            queryset = queryset.filter(
                Q(purchase_price__lte=purchase_price_max)
                | Q(
                    cost_layers__remaining_quantity__gt=0,
                    cost_layers__unit_cost__lte=purchase_price_max,
                )
            ).distinct()

        return queryset

    @action(
        detail=False,
        methods=['get'],
        url_path='pos',
        permission_classes=[IsAuthenticated, CanAccessPOS],
    )
    def pos(self, request):
        """List active sale-safe products without granting inventory access."""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @transaction.atomic
    def perform_create(self, serializer):
        product = serializer.save()
        AuditLog.log(
            user=self.request.user,
            action=AuditLog.ActionType.CREATE,
            model_name='Product',
            object_id=product.pk,
            object_repr=str(product),
            changes={'after': _product_audit_snapshot(product)},
            request=self.request,
        )

    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            product = get_object_or_404(
                Product.objects.select_for_update(),
                pk=kwargs['pk'],
            )
            self.check_object_permissions(request, product)
            before = _product_audit_snapshot(product)
            product.active = False
            product.save(update_fields=['active', 'updated_at'])
            AuditLog.log(
                user=request.user,
                action=AuditLog.ActionType.DELETE,
                model_name='Product',
                object_id=product.pk,
                object_repr=str(product),
                changes={
                    'before': before,
                    'after': _product_audit_snapshot(product),
                    'soft_deactivation': True,
                },
                request=request,
            )
        return Response(status=status.HTTP_204_NO_CONTENT)

    def perform_update(self, serializer):
        with transaction.atomic():
            locked_product = Product.objects.select_for_update().get(
                pk=serializer.instance.pk
            )
            serializer.instance = locked_product
            before = _product_audit_snapshot(locked_product)
            old_purchase_price = locked_product.purchase_price
            old_sale_price = locked_product.sale_price_ht
            product = serializer.save()
            ProductCostLayer.reconcile_to_stock(product)
            purchase_price_changed = product.purchase_price != old_purchase_price
            sale_price_changed = product.sale_price_ht != old_sale_price
            if purchase_price_changed or sale_price_changed:
                active_layers = product.cost_layers.select_for_update().filter(
                    remaining_quantity__gt=0,
                )
                if sale_price_changed:
                    active_layers.exclude(
                        sale_price=product.sale_price_ht,
                    ).update(sale_price=product.sale_price_ht)
                PriceHistory.objects.create(
                    product=product,
                    old_purchase_price=old_purchase_price,
                    new_purchase_price=product.purchase_price,
                    old_sale_price=old_sale_price,
                    new_sale_price=product.sale_price_ht,
                    changed_by=self.request.user,
                    reason='Modification de la fiche produit',
                )
            ProductCostLayer.assert_matches_stock(product)
            AuditLog.log(
                user=self.request.user,
                action=AuditLog.ActionType.UPDATE,
                model_name='Product',
                object_id=product.pk,
                object_repr=str(product),
                changes={
                    'before': before,
                    'after': _product_audit_snapshot(product),
                    'fields': sorted(serializer.validated_data.keys()),
                },
                request=self.request,
            )

    def _update_cost_layer(self, product, layer, request):
        serializer = ProductCostLayerSerializer(
            layer,
            data=request.data,
            partial=True,
            context={'request': request},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        product.refresh_from_db()
        return Response({
            'layer': serializer.data,
            'product': ProductSerializer(product, context={'request': request}).data,
        })

    @extend_schema(
        request=ProductCostLayerSerializer,
        responses={
            200: ProductCostLayerUpdateResponseSerializer,
            400: OpenApiResponse(description='Données de lot invalides.'),
            404: OpenApiResponse(description='Produit ou lot FIFO actif introuvable.'),
        },
        parameters=[
            OpenApiParameter(
                name='layer_id',
                type=OpenApiTypes.INT,
                location=OpenApiParameter.PATH,
                description='Identifiant du lot FIFO actif.',
            ),
        ],
    )
    @action(
        detail=True,
        methods=['patch'],
        url_path=r'cost-layers/(?P<layer_id>[^/.]+)',
        permission_classes=[IsAuthenticated, IsAdminRole],
    )
    def update_cost_layer(self, request, pk=None, layer_id=None):
        """Corriger un lot FIFO actif appartenant au produit."""
        product = self.get_object()
        try:
            layer = product.cost_layers.get(
                id=layer_id,
                remaining_quantity__gt=0,
            )
        except (ProductCostLayer.DoesNotExist, TypeError, ValueError):
            return Response({'detail': 'Lot FIFO introuvable pour ce produit.'}, status=404)
        return update_product_cost_layer_for_request(
            request,
            product.pk,
            layer_id_override=layer.pk,
        )

    @action(
        detail=True,
        methods=['patch'],
        url_path=r'cost-layers/by-position/(?P<position>[0-9]+)',
        permission_classes=[IsAuthenticated, IsAdminRole],
    )
    def update_cost_layer_by_position(self, request, pk=None, position=None):
        """Fallback: corriger un lot FIFO actif par sa position dans la liste."""
        request.data['index'] = position
        return update_product_cost_layer_for_request(request, pk)

    @action(
        detail=False,
        methods=['get'],
        permission_classes=[IsAuthenticated, IsAdminRole],
    )
    def stats(self, request):
        """Statistiques globales des produits"""
        products = self.get_queryset()

        total_products = products.count()
        active_products = products.filter(active=True).count()
        low_stock_count = products.filter(stock__lte=F('min_stock')).count()
        out_of_stock = products.filter(stock=0).count()

        valuation_products = products.prefetch_related(None).prefetch_related(Prefetch(
            'cost_layers',
            queryset=ProductCostLayer.objects.filter(
                remaining_quantity__gt=0,
            ).order_by('created_at', 'id'),
            to_attr='_stock_value_layers',
        ))
        stock_value = sum(
            (product.stock_value for product in valuation_products),
            Decimal('0.00'),
        )

        return Response({
            'total_products': total_products,
            'active_products': active_products,
            'low_stock_count': low_stock_count,
            'out_of_stock': out_of_stock,
            'stock_value': float(stock_value)
        })

    @action(
        detail=False,
        methods=['post'],
        parser_classes=[MultiPartParser],
        permission_classes=[IsAuthenticated, IsAdminRole],
        throttle_classes=[ProductImportRateThrottle],
    )
    def import_excel(self, request):
        """Import products from Excel/CSV file"""
        return self._import_products_file(request)

    def _normalize_import_column(self, value):
        text = unicodedata.normalize('NFKD', str(value or ''))
        text = text.encode('ascii', 'ignore').decode('ascii')
        text = text.lower().strip().replace('_', ' ').replace('-', ' ')
        text = ' '.join(text.split())
        column_mapping = {
            'nom': 'name',
            'name': 'name',
            'designation': 'name',
            'libelle': 'name',
            'titre': 'name',
            'produit': 'name',
            'code': 'barcode',
            'code barre': 'barcode',
            'barcode': 'barcode',
            'ean': 'barcode',
            'ref': 'barcode',
            'reference': 'barcode',
            'prix achat': 'purchase_price',
            'cout': 'purchase_price',
            'pa': 'purchase_price',
            'purchase price': 'purchase_price',
            'prix vente': 'sale_price',
            'prix vente ht': 'sale_price',
            'pv': 'sale_price',
            'prix': 'sale_price',
            'sale price': 'sale_price',
            'sale price ht': 'sale_price',
            'quantite': 'stock',
            'qte': 'stock',
            'qty': 'stock',
            'stock': 'stock',
            'min': 'min_stock',
            'seuil': 'min_stock',
            'sueil': 'min_stock',
            'alerte': 'min_stock',
            'stock min': 'min_stock',
            'min stock': 'min_stock',
            'categorie': 'category',
            'category': 'category',
            'famille': 'category',
            'rayon': 'category',
            'fournisseur': 'supplier',
            'supplier': 'supplier',
            'description': 'description',
            'tva': 'tva',
            'vat': 'tva',
            'image': 'image',
            'photo': 'image',
        }
        return column_mapping.get(text, text)

    def _is_blank_import_value(self, value):
        if value is None:
            return True
        if isinstance(value, str):
            return value.strip().lower() in {'', 'nan', 'none', 'null'}
        return False

    def _import_text(self, value, default=''):
        if self._is_blank_import_value(value):
            return default
        return str(value).strip()

    def _import_decimal(self, value, default='0'):
        if self._is_blank_import_value(value):
            return Decimal(default)
        try:
            return Decimal(str(value).strip().replace(',', '.'))
        except (InvalidOperation, ValueError) as exc:
            raise ValueError(f'Nombre invalide: {value}') from exc

    def _import_int(self, value, default=0):
        if self._is_blank_import_value(value):
            return default
        try:
            decimal_value = self._import_decimal(value, str(default))
            if decimal_value != decimal_value.to_integral_value():
                raise ValueError(f'Entier invalide: {value}')
            return int(decimal_value)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError(f'Entier invalide: {value}') from exc

    def _rows_from_csv_bytes(self, data):
        try:
            text = data.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = data.decode('cp1252')

        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;')
        except csv.Error:
            dialect = csv.excel

        reader = csv.DictReader(StringIO(text), dialect=dialect, strict=True)
        columns = [
            self._normalize_import_column(column)
            for column in (reader.fieldnames or [])
            if column
        ]
        rows = []
        for line_number, row in enumerate(reader, start=2):
            if len(rows) >= MAX_IMPORT_ROWS:
                raise ValueError(
                    f'Import limité à {MAX_IMPORT_ROWS} lignes.'
                )
            normalized = {}
            for key, value in row.items():
                if key is not None:
                    normalized[self._normalize_import_column(key)] = value
            rows.append((line_number, normalized))
        return columns, rows

    def _rows_from_excel_bytes(self, data):
        from openpyxl import load_workbook

        self._validate_zip_bytes(data, is_container=False)
        try:
            workbook = load_workbook(
                BytesIO(data),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as exc:
            raise ValueError('Fichier XLSX invalide ou corrompu.') from exc
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            return [], []

        columns = [self._normalize_import_column(header) for header in headers]
        rows = []
        for line_number, values in enumerate(iterator, start=2):
            if len(rows) >= MAX_IMPORT_ROWS:
                workbook.close()
                raise ValueError(
                    f'Import limité à {MAX_IMPORT_ROWS} lignes.'
                )
            normalized = {}
            for index, value in enumerate(values):
                if index < len(columns) and columns[index]:
                    normalized[columns[index]] = value
            rows.append((line_number, normalized))
        workbook.close()
        return columns, rows

    def _validate_zip_bytes(self, data, is_container=True):
        try:
            with zipfile.ZipFile(BytesIO(data)) as archive:
                infos = archive.infolist()
                if len(infos) > MAX_ZIP_MEMBERS:
                    raise ValueError(
                        f'Archive limitée à {MAX_ZIP_MEMBERS} fichiers.'
                    )
                total_size = 0
                for info in infos:
                    path = PurePosixPath(info.filename.replace('\\', '/'))
                    if path.is_absolute() or '..' in path.parts:
                        raise ValueError('Chemin ZIP non sûr détecté.')
                    if info.flag_bits & 0x1:
                        raise ValueError('Les archives chiffrées ne sont pas acceptées.')
                    total_size += info.file_size
                    if total_size > MAX_ZIP_UNCOMPRESSED_SIZE:
                        raise ValueError(
                            'Archive décompressée trop volumineuse '
                            f'(max {MAX_ZIP_UNCOMPRESSED_SIZE // (1024 * 1024)} Mo).'
                        )
                    if (
                        is_container
                        and info.file_size > 1024 * 1024
                        and info.compress_size > 0
                        and info.file_size / info.compress_size > 200
                    ):
                        raise ValueError('Taux de compression ZIP suspect.')
        except zipfile.BadZipFile:
            if is_container:
                raise
            raise ValueError('Fichier XLSX invalide.')

    def _rows_from_import_bytes(self, filename, data):
        lower_name = filename.lower()
        try:
            if lower_name.endswith('.csv'):
                return self._rows_from_csv_bytes(data)
            if lower_name.endswith(('.xlsx', '.xlsm')):
                return self._rows_from_excel_bytes(data)
            raise ValueError('Format non supporte. Utilise CSV, XLSX ou ZIP.')
        except ValueError:
            raise
        except (csv.Error, UnicodeError, zipfile.BadZipFile) as exc:
            raise ValueError('Fichier import invalide ou corrompu.') from exc

    def _zip_product_file(self, archive):
        preferred_names = (
            'produits.csv',
            'products.csv',
            'produits.xlsx',
            'products.xlsx',
        )
        members_by_name = {
            member.lower().replace('\\', '/').rsplit('/', 1)[-1]: member
            for member in archive.namelist()
        }
        for preferred_name in preferred_names:
            if preferred_name in members_by_name:
                return members_by_name[preferred_name]

        for member in archive.namelist():
            lower_member = member.lower()
            if lower_member.endswith(('.csv', '.xlsx', '.xlsm')):
                return member
        return None

    def _zip_image_members(self, archive):
        allowed_extensions = ('.jpg', '.jpeg', '.png', '.webp', '.gif')
        by_stem = {}
        by_name = {}
        for member in archive.namelist():
            lower_member = member.lower()
            if not lower_member.endswith(allowed_extensions):
                continue
            name = lower_member.rsplit('/', 1)[-1]
            stem = name.rsplit('.', 1)[0]
            by_name[name] = member
            by_stem.setdefault(stem, member)
        return by_name, by_stem

    def _validated_zip_image(self, archive, member):
        info = archive.getinfo(member)
        if info.file_size > MAX_IMPORT_IMAGE_SIZE:
            raise ValueError(
                f'Image {member} trop volumineuse (max 2 Mo).'
            )
        data = archive.read(member)
        try:
            from PIL import Image

            with Image.open(BytesIO(data)) as image:
                width, height = image.size
                image_pixels = width * height
                image.verify()
                image_format = (image.format or '').upper()
        except Exception as exc:
            raise ValueError(f'Image invalide: {member}.') from exc
        if image_pixels > MAX_IMPORT_IMAGE_PIXELS:
            raise ValueError(
                f'Image {member} trop grande '
                f'(max {MAX_IMPORT_IMAGE_PIXELS} pixels).'
            )
        if image_format not in ALLOWED_IMPORT_IMAGE_FORMATS:
            raise ValueError(f'Format image non supporté: {member}.')
        return ContentFile(data)

    def _zip_image_for_row(self, archive, image_maps, barcode, image_ref):
        if not archive:
            return None, None

        by_name, by_stem = image_maps
        candidates = []
        image_text = self._import_text(image_ref)
        if image_text:
            normalized = image_text.replace('\\', '/').lower().rsplit('/', 1)[-1]
            candidates.append(normalized)
            candidates.append(normalized.rsplit('.', 1)[0])
        if barcode:
            candidates.append(str(barcode).lower())

        for candidate in candidates:
            member = by_name.get(candidate) or by_stem.get(candidate)
            if member:
                filename = member.rsplit('/', 1)[-1]
                return filename, self._validated_zip_image(archive, member)
        return None, None

    def _import_products_file(self, request):
        """Import products from CSV, XLSX or ZIP without pandas."""
        if 'file' not in request.FILES:
            return Response(
                {'detail': 'Aucun fichier fourni.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        uploaded_file = request.FILES['file']
        if uploaded_file.size > MAX_IMPORT_FILE_SIZE:
            return Response(
                {
                    'detail': (
                        'Fichier trop volumineux '
                        f'(max {MAX_IMPORT_FILE_SIZE // (1024 * 1024)} Mo).'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        dry_run = str(request.data.get('dry_run', '')).lower() in {
            '1', 'true', 'yes', 'on'
        }
        upsert = str(request.data.get('upsert', '')).lower() in {
            '1', 'true', 'yes', 'on'
        }
        archive = None
        image_maps = ({}, {})
        saved_images = []

        def cleanup_saved_images():
            for storage, name in saved_images:
                try:
                    storage.delete(name)
                except Exception:
                    logger.warning('Unable to clean imported image %s', name)

        try:
            uploaded_data = uploaded_file.read()
            source_name = uploaded_file.name

            if source_name.lower().endswith('.zip'):
                self._validate_zip_bytes(uploaded_data)
                archive = zipfile.ZipFile(BytesIO(uploaded_data))
                product_member = self._zip_product_file(archive)
                if not product_member:
                    return Response(
                        {'detail': 'Le ZIP doit contenir un fichier CSV ou XLSX.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                source_name = product_member
                uploaded_data = archive.read(product_member)
                image_maps = self._zip_image_members(archive)

            columns, rows = self._rows_from_import_bytes(source_name, uploaded_data)
            if 'name' not in columns or 'barcode' not in columns:
                found_cols = ', '.join(columns)
                return Response(
                    {
                        'detail': (
                            'Colonnes obligatoires manquantes : "name" '
                            '(ou Nom) et "barcode" (ou EAN/Code). '
                            f'Colonnes trouvees : {found_cols}'
                        ),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            prepared_rows = []
            errors = []
            seen_barcodes = set()

            for line_number, row in rows:
                try:
                    barcode = self._import_text(row.get('barcode'))
                    name = self._import_text(row.get('name'))
                    if not barcode or not name:
                        raise ValueError('Nom et code-barres obligatoires.')
                    if barcode in seen_barcodes:
                        raise ValueError(
                            f'Code-barres dupliqué dans le fichier: {barcode}.'
                        )
                    seen_barcodes.add(barcode)

                    row_payload = {
                        'name': name,
                        'barcode': barcode,
                        'description': self._import_text(row.get('description')),
                        'purchase_price': self._import_decimal(
                            row.get('purchase_price'), '0'
                        ),
                        'sale_price': self._import_decimal(
                            row.get('sale_price'), '0'
                        ),
                        'tva': self._import_decimal(row.get('tva'), '0'),
                        'stock': self._import_int(row.get('stock'), 0),
                        'min_stock': self._import_int(row.get('min_stock'), 5),
                        'category': self._import_text(row.get('category'), 'General'),
                        'supplier': self._import_text(row.get('supplier')),
                        'image': self._import_text(row.get('image')),
                    }
                    provided_fields = {
                        field
                        for field in (
                            'description', 'purchase_price', 'sale_price',
                            'tva', 'stock', 'min_stock', 'category',
                            'supplier', 'image',
                        )
                        if field in row
                        and not self._is_blank_import_value(row.get(field))
                    }
                    row_serializer = ProductImportRowSerializer(data=row_payload)
                    row_serializer.is_valid(raise_exception=True)
                    image_name, image_content = self._zip_image_for_row(
                        archive,
                        image_maps,
                        barcode,
                        row.get('image'),
                    )
                    prepared_rows.append({
                        'line': line_number,
                        'data': row_serializer.validated_data,
                        'provided_fields': provided_fields,
                        'image_name': image_name,
                        'image_content': image_content,
                    })
                except Exception as exc:
                    if len(errors) < MAX_IMPORT_ERRORS:
                        detail = getattr(exc, 'detail', exc)
                        errors.append(f'Ligne {line_number}: {detail}')

            if errors:
                return Response(
                    {
                        'detail': 'Import annulé: aucune donnée n’a été modifiée.',
                        'errors': errors,
                        'errors_truncated': len(errors) >= MAX_IMPORT_ERRORS,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not rows:
                return Response(
                    {'detail': 'Le fichier ne contient aucune ligne produit.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            existing_barcodes = set(
                Product.objects.filter(barcode__in=seen_barcodes)
                .values_list('barcode', flat=True)
            )
            would_create = len(seen_barcodes - existing_barcodes)
            would_update = len(existing_barcodes) if upsert else 0
            would_skip = len(existing_barcodes) if not upsert else 0
            if dry_run:
                return Response({
                    'dry_run': True,
                    'valid_rows': len(prepared_rows),
                    'would_create': would_create,
                    'would_update': would_update,
                    'would_skip': would_skip,
                    'errors': [],
                })

            created_count = 0
            updated_count = 0
            images_count = 0
            skipped_count = 0
            with transaction.atomic():
                existing_products = {
                    product.barcode: product
                    for product in Product.objects.select_for_update()
                    .filter(barcode__in=seen_barcodes)
                    .order_by('pk')
                }

                for prepared in prepared_rows:
                    data = prepared['data']
                    provided_fields = prepared['provided_fields']
                    barcode = data['barcode']
                    product = existing_products.get(barcode)
                    if product and not upsert:
                        skipped_count += 1
                        continue

                    if product:
                        audit_action = AuditLog.ActionType.UPDATE
                        audit_before = _product_audit_snapshot(product)
                        ProductCostLayer.reconcile_to_stock(
                            product,
                            note='Réconciliation avant import',
                        )
                        old_purchase_price = product.purchase_price
                        old_sale_price = product.sale_price_ht
                        target_stock = (
                            data['stock']
                            if 'stock' in provided_fields
                            else product.stock
                        )
                        product.name = data['name']
                        update_fields = ['name']
                        for source_field, model_field in (
                            ('description', 'description'),
                            ('purchase_price', 'purchase_price'),
                            ('sale_price', 'sale_price_ht'),
                            ('tva', 'tva'),
                            ('min_stock', 'min_stock'),
                        ):
                            if source_field in provided_fields:
                                setattr(product, model_field, data[source_field])
                                update_fields.append(model_field)
                        if 'category' in provided_fields:
                            product.category, _ = Category.objects.get_or_create(
                                name=data['category'],
                                defaults={
                                    'description': 'Auto-created from import'
                                },
                            )
                            update_fields.append('category')
                        if 'supplier' in provided_fields:
                            product.supplier, _ = Supplier.objects.get_or_create(
                                name=data['supplier'],
                            )
                            update_fields.append('supplier')
                        product.save(update_fields=[*update_fields, 'updated_at'])
                        if target_stock != product.stock:
                            StockMovement.objects.create(
                                product=product,
                                movement_type=StockMovement.MovementType.ADJUST,
                                quantity=target_stock,
                                notes=f'Import produits ligne {prepared["line"]}',
                                created_by=request.user,
                            )
                            product.refresh_from_db(fields=['stock'])
                        if (
                            old_purchase_price != product.purchase_price
                            or old_sale_price != product.sale_price_ht
                        ):
                            PriceHistory.objects.create(
                                product=product,
                                old_purchase_price=old_purchase_price,
                                new_purchase_price=product.purchase_price,
                                old_sale_price=old_sale_price,
                                new_sale_price=product.sale_price_ht,
                                changed_by=request.user,
                                reason='Mise à jour par import',
                            )
                        updated_count += 1
                    else:
                        audit_action = AuditLog.ActionType.CREATE
                        audit_before = None
                        category, _ = Category.objects.get_or_create(
                            name=data['category'] or 'General',
                            defaults={
                                'description': 'Auto-created from import'
                            },
                        )
                        supplier = None
                        if data['supplier']:
                            supplier, _ = Supplier.objects.get_or_create(
                                name=data['supplier'],
                            )
                        imported_stock = data['stock']
                        product = Product.objects.create(
                            name=data['name'],
                            barcode=barcode,
                            description=data['description'],
                            purchase_price=data['purchase_price'],
                            sale_price_ht=data['sale_price'],
                            tva=data['tva'],
                            stock=0,
                            min_stock=data['min_stock'],
                            category=category,
                            supplier=supplier,
                        )
                        if imported_stock > 0:
                            StockMovement.objects.create(
                                product=product,
                                movement_type=StockMovement.MovementType.IN,
                                quantity=imported_stock,
                                unit_cost=product.purchase_price,
                                sale_price=product.sale_price_ht,
                                supplier=product.supplier,
                                reference=f'IMPORT-LIGNE-{prepared["line"]}',
                                notes='Stock initial créé par import produits',
                                created_by=request.user,
                            )
                            product.refresh_from_db(fields=['stock'])
                        created_count += 1

                    if prepared['image_name'] and prepared['image_content']:
                        product.image.save(
                            prepared['image_name'],
                            prepared['image_content'],
                            save=False,
                        )
                        saved_images.append((product.image.storage, product.image.name))
                        product.save(update_fields=['image', 'updated_at'])
                        images_count += 1
                    ProductCostLayer.reconcile_to_stock(product)
                    ProductCostLayer.assert_matches_stock(product)
                    audit_changes = {
                        'after': _product_audit_snapshot(product),
                        'import_line': prepared['line'],
                        'fields': sorted({'name', *provided_fields}),
                    }
                    if audit_before is not None:
                        audit_changes['before'] = audit_before
                    AuditLog.log(
                        user=request.user,
                        action=audit_action,
                        model_name='Product',
                        object_id=product.pk,
                        object_repr=str(product),
                        changes=audit_changes,
                        request=request,
                    )

            return Response(
                {
                    'created': created_count,
                    'updated': updated_count,
                    'images': images_count,
                    'skipped': skipped_count,
                    'errors': errors,
                    'dry_run': False,
                    'upsert': upsert,
                },
                status=(
                    status.HTTP_201_CREATED
                    if created_count
                    else status.HTTP_200_OK
                ),
            )
        except ValueError as exc:
            cleanup_saved_images()
            return Response(
                {'detail': str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except zipfile.BadZipFile:
            cleanup_saved_images()
            return Response(
                {'detail': 'Fichier ZIP invalide.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except IntegrityError:
            cleanup_saved_images()
            logger.info('Product import stopped by a concurrent data conflict')
            return Response(
                {
                    'detail': (
                        'Import annulé à cause d’un conflit de données. '
                        'Recharge les produits puis réessaie.'
                    )
                },
                status=status.HTTP_409_CONFLICT,
            )
        except Exception:
            cleanup_saved_images()
            logger.exception("Product import failed")
            return Response(
                {'detail': "Erreur interne lors de l'import."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        finally:
            if archive:
                archive.close()

    @action(detail=True, methods=['post'])
    def add_stock(self, request, pk=None):
        """Ajouter du stock à un produit"""
        product = self.get_object()
        serializer = StockInSerializer(data={
            'product': product.id,
            **request.data
        }, context={'request': request})

        if serializer.is_valid():
            movement = serializer.save()
            product.refresh_from_db(fields=['stock'])
            return Response({
                'message': f'{movement.quantity} unités ajoutées au stock',
                'new_stock': product.stock,
                'movement_id': movement.id
            })
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class StockMovementViewSet(viewsets.ModelViewSet):
    """API pour les mouvements de stock"""
    queryset = StockMovement.objects.select_related(
        'product', 'supplier', 'created_by'
    ).all()
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated, CanViewInventory, CanManageInventory]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['product', 'movement_type', 'supplier']
    ordering_fields = ['created_at']
    ordering = ['-created_at']
    http_method_names = ['get', 'post', 'head', 'options']  # Pas de modification/suppression

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filtre par période
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        return queryset

    @action(detail=False, methods=['post'])
    def stock_in(self, request):
        """Entrée de stock (réapprovisionnement)"""
        serializer = StockInSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            movement = serializer.save()
            return Response(
                StockMovementSerializer(
                    movement,
                    context={'request': request},
                ).data,
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_stock_in(self, request):
        """Create a bounded, fully validated batch with no partial writes."""
        serializer = BulkStockInSerializer(
            data=request.data,
            context={'request': request},
        )
        serializer.is_valid(raise_exception=True)
        movements = serializer.save()
        results = [
            {
                'product_id': movement.product_id,
                'quantity': movement.quantity,
                'success': True,
                'movement_id': movement.id,
            }
            for movement in movements
        ]

        return Response({
            'success': results,
            'errors': [],
            'total_success': len(results),
            'total_errors': 0,
        })


class ProductCostLayerViewSet(viewsets.ModelViewSet):
    """API pour consulter et corriger les lots FIFO actifs."""
    queryset = ProductCostLayer.objects.select_related('product').all()
    serializer_class = ProductCostLayerSerializer
    permission_classes = [IsAuthenticated, IsAdminRole]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['product']
    ordering_fields = ['created_at', 'unit_cost', 'sale_price']
    ordering = ['created_at', 'id']
    http_method_names = ['get', 'patch', 'head', 'options']

    def get_queryset(self):
        queryset = super().get_queryset()
        active_only = self.request.query_params.get('active_only')
        if active_only and active_only.lower() == 'true':
            queryset = queryset.filter(remaining_quantity__gt=0)
        return queryset

    def update(self, request, *args, **kwargs):
        layer = self.get_object()
        return update_product_cost_layer_for_request(
            request,
            layer.product_id,
            layer_id_override=layer.pk,
        )


class PurchaseOrderViewSet(viewsets.ModelViewSet):
    """API pour les commandes fournisseurs"""
    queryset = (
        PurchaseOrder.objects.select_related('supplier', 'created_by')
        .prefetch_related(
            'items__product__cost_layers',
            'payments__created_by',
            'payments__reversed_by',
        )
        .all()
    )
    permission_classes = [IsAuthenticated, IsAdminRole]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['supplier', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'create':
            return PurchaseOrderCreateSerializer
        return PurchaseOrderSerializer

    def _payment_operation_id(self, request, validated_data):
        operation_id = (
            validated_data.get('operation_id')
            or request.headers.get('Idempotency-Key')
        )
        if not operation_id:
            return None
        if not re.fullmatch(r'[A-Za-z0-9._:-]{8,64}', operation_id):
            return None
        return operation_id

    def _payment_payload_hash(self, payload):
        canonical = json.dumps(
            payload,
            ensure_ascii=False,
            sort_keys=True,
            separators=(',', ':'),
        ).encode('utf-8')
        return hashlib.sha256(canonical).hexdigest()

    def _payment_response(self, request, order_id, payment, *, replay, http_status):
        order = self.get_queryset().get(pk=order_id)
        return Response(
            {
                'payment': SupplierPaymentSerializer(payment).data,
                'order': PurchaseOrderSerializer(
                    order,
                    context={'request': request},
                ).data,
                'idempotent_replay': replay,
            },
            status=http_status,
        )

    @action(detail=True, methods=['post'], url_path='payments')
    def payments(self, request, pk=None):
        """Enregistre un règlement fournisseur sans affecter le résultat."""
        input_serializer = SupplierPaymentCreateSerializer(data=request.data)
        input_serializer.is_valid(raise_exception=True)
        payment_data = input_serializer.validated_data
        operation_id = self._payment_operation_id(request, payment_data)
        if not operation_id:
            return Response(
                {
                    'operation_id': [
                        'Un identifiant idempotent valide (8 à 64 caractères) est obligatoire.'
                    ]
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        reference = payment_data.get('reference', '').strip()
        note = payment_data.get('note', '').strip()
        payload_hash = self._payment_payload_hash({
            'order_id': int(pk),
            'amount': format(payment_data['amount'], 'f'),
            'method': payment_data['method'],
            'paid_on': payment_data['paid_on'].isoformat(),
            'reference': reference,
            'note': note,
        })

        with transaction.atomic():
            order = get_object_or_404(
                PurchaseOrder.objects.select_for_update(),
                pk=pk,
            )
            self.check_object_permissions(request, order)

            existing = SupplierPayment.objects.select_for_update().filter(
                operation_id=operation_id,
            ).first()
            if existing:
                if (
                    existing.order_id != order.id
                    or existing.operation_payload_hash != payload_hash
                ):
                    return Response(
                        {
                            'operation_id': [
                                'Cet identifiant a déjà été utilisé avec un autre règlement.'
                            ]
                        },
                        status=status.HTTP_409_CONFLICT,
                    )
                return self._payment_response(
                    request,
                    order.id,
                    existing,
                    replay=True,
                    http_status=status.HTTP_200_OK,
                )

            if order.status in {
                PurchaseOrder.OrderStatus.DRAFT,
                PurchaseOrder.OrderStatus.CANCELLED,
            }:
                return Response(
                    {
                        'detail': (
                            'Une commande brouillon ou annulée ne peut pas être réglée.'
                        )
                    },
                    status=status.HTTP_409_CONFLICT,
                )

            # Le verrou de commande sérialise deux règlements simultanés et
            # garantit que le contrôle de solde et la création sont atomiques.
            total_amount = sum(
                (item.total for item in order.items.all()),
                Decimal('0.00'),
            )
            paid_amount = (
                SupplierPayment.objects.select_for_update().filter(
                    order=order,
                    status=SupplierPayment.PaymentStatus.ACTIVE,
                ).aggregate(total=Sum('amount'))['total']
                or Decimal('0.00')
            )
            balance_due = total_amount - paid_amount
            if payment_data['amount'] > balance_due:
                return Response(
                    {
                        'amount': [
                            f'Le règlement dépasse le solde restant de {balance_due:.2f} DH.'
                        ]
                    },
                    status=status.HTTP_409_CONFLICT,
                )

            try:
                # Savepoint local : une collision très rare entre deux
                # commandes ne casse pas la transaction ni l'idempotence.
                with transaction.atomic():
                    payment = SupplierPayment.objects.create(
                        order=order,
                        amount=payment_data['amount'],
                        method=payment_data['method'],
                        paid_on=payment_data['paid_on'],
                        reference=reference,
                        note=note,
                        created_by=request.user,
                        operation_id=operation_id,
                        operation_payload_hash=payload_hash,
                    )
            except IntegrityError:
                existing = SupplierPayment.objects.filter(
                    operation_id=operation_id,
                ).first()
                if existing and (
                    existing.order_id == order.id
                    and existing.operation_payload_hash == payload_hash
                ):
                    return self._payment_response(
                        request,
                        order.id,
                        existing,
                        replay=True,
                        http_status=status.HTTP_200_OK,
                    )
                if existing:
                    return Response(
                        {
                            'operation_id': [
                                'Cet identifiant a déjà été utilisé avec un autre règlement.'
                            ]
                        },
                        status=status.HTTP_409_CONFLICT,
                    )
                raise
            AuditLog.log(
                user=request.user,
                action=AuditLog.ActionType.CREATE,
                model_name='SupplierPayment',
                object_id=payment.id,
                object_repr=(
                    f'{order.reference}: {payment.amount} {payment.method}'
                ),
                changes={
                    'order_id': order.id,
                    'order_reference': order.reference,
                    'amount': str(payment.amount),
                    'method': payment.method,
                    'paid_on': payment.paid_on.isoformat(),
                    'reference': payment.reference,
                },
                request=request,
            )

        return self._payment_response(
            request,
            order.id,
            payment,
            replay=False,
            http_status=status.HTTP_201_CREATED,
        )

    @action(
        detail=True,
        methods=['post'],
        url_path=r'payments/(?P<payment_id>\d+)/reverse',
    )
    def reverse_payment(self, request, pk=None, payment_id=None):
        """Contrepasse durablement un règlement ; aucune suppression physique."""
        input_serializer = SupplierPaymentReverseSerializer(data=request.data)
        input_serializer.is_valid(raise_exception=True)
        reversal_data = input_serializer.validated_data
        operation_id = self._payment_operation_id(request, reversal_data)
        if not operation_id:
            return Response(
                {
                    'operation_id': [
                        'Un identifiant idempotent valide (8 à 64 caractères) est obligatoire.'
                    ]
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        reason = reversal_data['reason'].strip()
        payload_hash = self._payment_payload_hash({
            'order_id': int(pk),
            'payment_id': int(payment_id),
            'reason': reason,
        })

        with transaction.atomic():
            order = get_object_or_404(
                PurchaseOrder.objects.select_for_update(),
                pk=pk,
            )
            self.check_object_permissions(request, order)
            payment = get_object_or_404(
                SupplierPayment.objects.select_for_update(),
                pk=payment_id,
                order=order,
            )

            reused_operation = SupplierPayment.objects.select_for_update().filter(
                reversal_operation_id=operation_id,
            ).exclude(pk=payment.pk).exists()
            if reused_operation:
                return Response(
                    {
                        'operation_id': [
                            'Cet identifiant a déjà été utilisé pour une autre contrepassation.'
                        ]
                    },
                    status=status.HTTP_409_CONFLICT,
                )

            if payment.status == SupplierPayment.PaymentStatus.REVERSED:
                if (
                    payment.reversal_operation_id == operation_id
                    and payment.reversal_payload_hash == payload_hash
                ):
                    return self._payment_response(
                        request,
                        order.id,
                        payment,
                        replay=True,
                        http_status=status.HTTP_200_OK,
                    )
                return Response(
                    {'detail': 'Ce règlement a déjà été contrepassé.'},
                    status=status.HTTP_409_CONFLICT,
                )

            payment.status = SupplierPayment.PaymentStatus.REVERSED
            payment.reversed_by = request.user
            payment.reversed_at = timezone.now()
            payment.reversal_reason = reason
            payment.reversal_operation_id = operation_id
            payment.reversal_payload_hash = payload_hash
            payment.save(update_fields=[
                'status', 'reversed_by', 'reversed_at', 'reversal_reason',
                'reversal_operation_id', 'reversal_payload_hash',
            ])
            AuditLog.log(
                user=request.user,
                action=AuditLog.ActionType.UPDATE,
                model_name='SupplierPayment',
                object_id=payment.id,
                object_repr=f'Contrepassation {order.reference}: {payment.amount}',
                changes={
                    'status': {
                        'before': SupplierPayment.PaymentStatus.ACTIVE,
                        'after': SupplierPayment.PaymentStatus.REVERSED,
                    },
                    'reason': reason,
                },
                request=request,
            )

        return self._payment_response(
            request,
            order.id,
            payment,
            replay=False,
            http_status=status.HTTP_200_OK,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        with transaction.atomic():
            order = get_object_or_404(
                PurchaseOrder.objects.select_for_update(),
                pk=kwargs['pk'],
            )
            self.check_object_permissions(request, order)
            if order.status != PurchaseOrder.OrderStatus.DRAFT:
                return Response(
                    {'detail': 'Seule une commande brouillon peut être modifiée.'},
                    status=status.HTTP_409_CONFLICT,
                )
            serializer = self.get_serializer(
                order,
                data=request.data,
                partial=partial,
            )
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            order = get_object_or_404(
                PurchaseOrder.objects.select_for_update(),
                pk=kwargs['pk'],
            )
            self.check_object_permissions(request, order)
            if order.status not in {
                PurchaseOrder.OrderStatus.DRAFT,
                PurchaseOrder.OrderStatus.CANCELLED,
            }:
                return Response(
                    {
                        'detail': (
                            'Une commande envoyée ou réceptionnée '
                            'ne peut pas être supprimée.'
                        )
                    },
                    status=status.HTTP_409_CONFLICT,
                )
            if order.payments.exists():
                return Response(
                    {
                        'detail': (
                            'Une commande ayant un historique de règlements '
                            'ne peut pas être supprimée.'
                        )
                    },
                    status=status.HTTP_409_CONFLICT,
                )
            self.perform_destroy(order)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Marquer commande comme envoyée"""
        with transaction.atomic():
            order = get_object_or_404(
                PurchaseOrder.objects.select_for_update(),
                pk=pk,
            )
            self.check_object_permissions(request, order)
            if order.status != PurchaseOrder.OrderStatus.DRAFT:
                return Response(
                    {'detail': 'Seule une commande brouillon peut être envoyée.'},
                    status=status.HTTP_409_CONFLICT,
                )
            if not order.items.exists():
                return Response(
                    {'detail': 'La commande doit contenir au moins une ligne.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            order.status = PurchaseOrder.OrderStatus.SENT
            order.save(update_fields=['status', 'updated_at'])
        return Response({'status': 'Commande envoyée'})

    @action(detail=True, methods=['post'])
    def receive(self, request, pk=None):
        """Receive validated order lines exactly once inside a locked transaction."""
        payload = PurchaseOrderReceiveSerializer(data=request.data)
        payload.is_valid(raise_exception=True)
        received_items = payload.validated_data['items']
        supplied_receipt_id = (
            payload.validated_data.get('receipt_id')
            or request.headers.get('Idempotency-Key')
        )
        if supplied_receipt_id and not re.fullmatch(
            r'[A-Za-z0-9._:-]{8,64}',
            supplied_receipt_id,
        ):
            return Response(
                {'receipt_id': 'Identifiant de réception invalide.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        receipt_id = supplied_receipt_id or f'legacy-{uuid.uuid4().hex}'
        canonical_items = []
        for entry in sorted(received_items, key=lambda value: value['item_id']):
            canonical_items.append({
                key: (str(value) if isinstance(value, Decimal) else value)
                for key, value in sorted(entry.items())
            })
        payload_hash = hashlib.sha256(
            json.dumps(
                canonical_items,
                ensure_ascii=False,
                sort_keys=True,
                separators=(',', ':'),
            ).encode('utf-8')
        ).hexdigest()

        with transaction.atomic():
            order = get_object_or_404(
                PurchaseOrder.objects.select_for_update().select_related(
                    'supplier'
                ),
                pk=pk,
            )
            self.check_object_permissions(request, order)
            previous_receipt = PurchaseReceipt.objects.select_for_update().filter(
                order=order,
                receipt_id=receipt_id,
            ).first()
            if previous_receipt:
                if previous_receipt.payload_hash != payload_hash:
                    return Response(
                        {
                            'detail': (
                                'Ce receipt_id a déjà été utilisé avec '
                                'un contenu différent.'
                            )
                        },
                        status=status.HTTP_409_CONFLICT,
                    )
                return Response({
                    'order': PurchaseOrderSerializer(
                        order,
                        context={'request': request},
                    ).data,
                    'results': previous_receipt.result.get('results', []),
                    'receipt_id': receipt_id,
                    'idempotent_replay': True,
                    'idempotency_protected': True,
                })
            if order.status not in {
                PurchaseOrder.OrderStatus.SENT,
                PurchaseOrder.OrderStatus.PARTIALLY_RECEIVED,
            }:
                return Response(
                    {'detail': 'Commande non envoyée — réception impossible.'},
                    status=status.HTTP_409_CONFLICT,
                )

            requested_ids = [entry['item_id'] for entry in received_items]
            locked_items = {
                item.pk: item
                for item in PurchaseOrderItem.objects.select_for_update()
                .filter(order=order, pk__in=requested_ids)
                .select_related('product')
                .order_by('pk')
            }
            missing_ids = sorted(set(requested_ids) - set(locked_items))
            if missing_ids:
                return Response(
                    {'items': f'Lignes inconnues pour cette commande: {missing_ids}.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            product_ids = sorted({item.product_id for item in locked_items.values()})
            products = {
                product.pk: product
                for product in Product.objects.select_for_update()
                .filter(pk__in=product_ids)
                .order_by('pk')
            }

            prepared = []
            for received in received_items:
                item = locked_items[received['item_id']]
                remaining = item.quantity - item.received_quantity
                if remaining <= 0:
                    return Response(
                        {'items': f'La ligne {item.pk} est déjà entièrement réceptionnée.'},
                        status=status.HTTP_409_CONFLICT,
                    )
                qty = received.get('quantity', remaining)
                if qty > remaining:
                    return Response(
                        {
                            'items': (
                                f'Sur-réception interdite pour la ligne {item.pk}: '
                                f'{remaining} unité(s) restante(s).'
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                prepared.append((item, products[item.product_id], qty, received))

            # Le montant payable utilise le coût réel de chaque quantité reçue
            # et le coût commandé pour le reliquat. Une hausse laisse donc un
            # nouveau solde à payer ; une baisse qui créerait un trop-perçu est
            # refusée avant tout mouvement de stock.
            projected_total = order.total_amount
            for item, _product, qty, received in prepared:
                applied_cost = received.get('unit_cost', item.unit_cost)
                projected_total += qty * (applied_cost - item.unit_cost)
            active_paid = (
                SupplierPayment.objects.select_for_update().filter(
                    order=order,
                    status=SupplierPayment.PaymentStatus.ACTIVE,
                ).aggregate(total=Sum('amount'))['total']
                or Decimal('0.00')
            )
            if active_paid > projected_total:
                return Response(
                    {
                        'items': (
                            'Le nouveau coût réel créerait un trop-perçu fournisseur. '
                            'Contrepassez d’abord un règlement concerné.'
                        )
                    },
                    status=status.HTTP_409_CONFLICT,
                )

            results = []
            for item, product, qty, received in prepared:
                applied_cost = received.get('unit_cost', item.unit_cost)
                new_sale_price = received.get('new_sale_price')
                should_update_sale_price = bool(
                    received.get('update_sale_price')
                    and new_sale_price is not None
                )
                applied_sale_price = (
                    new_sale_price
                    if should_update_sale_price
                    else product.sale_price_ht
                )

                old_purchase_price = product.purchase_price
                old_sale_price = product.sale_price_ht
                StockMovement.objects.create(
                    product=product,
                    movement_type=StockMovement.MovementType.IN,
                    quantity=qty,
                    unit_cost=applied_cost,
                    sale_price=applied_sale_price,
                    supplier=order.supplier,
                    reference=f'{order.reference}:{receipt_id}',
                    notes=f'Réception commande {order.reference}',
                    created_by=request.user,
                )
                item.received_quantity += qty
                item.received_cost_total += qty * applied_cost
                item.save(update_fields=[
                    'received_quantity', 'received_cost_total',
                ])

                product_changes = []
                if received.get('update_purchase_price'):
                    product.purchase_price = applied_cost
                    product_changes.append('purchase_price')
                # Fournir un nouveau prix signifie explicitement qu'il devient
                # le prix courant unique, pour l'ancien et le nouveau stock.
                if (
                    should_update_sale_price
                    and product.sale_price_ht != new_sale_price
                ):
                    product.sale_price_ht = new_sale_price
                    product_changes.append('sale_price_ht')
                if product_changes:
                    product_changes.append('updated_at')
                    product.save(update_fields=product_changes)
                    PriceHistory.objects.create(
                        product=product,
                        old_purchase_price=old_purchase_price,
                        new_purchase_price=product.purchase_price,
                        old_sale_price=old_sale_price,
                        new_sale_price=product.sale_price_ht,
                        changed_by=request.user,
                        reason=f'Réception commande {order.reference}',
                    )

                results.append({
                    'item_id': item.id,
                    'product': product.name,
                    'received': qty,
                    'unit_cost_applied': float(applied_cost),
                    'sale_price_applied': float(applied_sale_price),
                    'updated_purchase_price': bool(
                        received.get('update_purchase_price')
                    ),
                    'updated_sale_price': bool(
                        should_update_sale_price
                        and new_sale_price != old_sale_price
                    ),
                })

            all_received = not PurchaseOrderItem.objects.filter(
                order=order,
                received_quantity__lt=F('quantity'),
            ).exists()
            order.status = (
                PurchaseOrder.OrderStatus.RECEIVED
                if all_received
                else PurchaseOrder.OrderStatus.PARTIALLY_RECEIVED
            )
            order.save(update_fields=['status', 'updated_at'])
            order.refresh_from_db()
            PurchaseReceipt.objects.create(
                order=order,
                receipt_id=receipt_id,
                payload_hash=payload_hash,
                result={'results': results},
                created_by=request.user,
            )

        return Response({
            'order': PurchaseOrderSerializer(order, context={'request': request}).data,
            'results': results,
            'receipt_id': receipt_id,
            'idempotent_replay': False,
            'idempotency_protected': bool(supplied_receipt_id),
        })

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Annuler la commande"""
        with transaction.atomic():
            order = get_object_or_404(
                PurchaseOrder.objects.select_for_update(),
                pk=pk,
            )
            self.check_object_permissions(request, order)
            if order.status not in {
                PurchaseOrder.OrderStatus.DRAFT,
                PurchaseOrder.OrderStatus.SENT,
            }:
                return Response(
                    {'detail': 'Cette commande ne peut plus être annulée.'},
                    status=status.HTTP_409_CONFLICT,
                )
            if order.items.filter(received_quantity__gt=0).exists():
                return Response(
                    {'detail': 'Une commande partiellement réceptionnée doit être contrepassée.'},
                    status=status.HTTP_409_CONFLICT,
                )
            if order.payments.filter(
                status=SupplierPayment.PaymentStatus.ACTIVE,
            ).exists():
                return Response(
                    {
                        'detail': (
                            'Contrepassez les règlements actifs avant '
                            'd’annuler cette commande.'
                        )
                    },
                    status=status.HTTP_409_CONFLICT,
                )
            order.status = PurchaseOrder.OrderStatus.CANCELLED
            order.save(update_fields=['status', 'updated_at'])
        return Response({'status': 'Commande annulée'})


class InventoryCountViewSet(viewsets.ModelViewSet):
    """API pour les inventaires physiques"""
    queryset = InventoryCount.objects.select_related(
        'counted_by', 'validated_by'
    ).prefetch_related('items__product').all()
    permission_classes = [IsAuthenticated, IsAdminRole]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'create':
            return InventoryCountCreateSerializer
        return InventoryCountSerializer

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        with transaction.atomic():
            count = get_object_or_404(
                InventoryCount.objects.select_for_update(),
                pk=kwargs['pk'],
            )
            self.check_object_permissions(request, count)
            if count.status != InventoryCount.CountStatus.IN_PROGRESS:
                return Response(
                    {'detail': 'Un inventaire terminé ne peut plus être modifié.'},
                    status=status.HTTP_409_CONFLICT,
                )
            serializer = self.get_serializer(
                count,
                data=request.data,
                partial=partial,
            )
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        with transaction.atomic():
            count = get_object_or_404(
                InventoryCount.objects.select_for_update(),
                pk=kwargs['pk'],
            )
            self.check_object_permissions(request, count)
            if count.status != InventoryCount.CountStatus.IN_PROGRESS:
                return Response(
                    {'detail': 'Un inventaire terminé ne peut pas être supprimé.'},
                    status=status.HTTP_409_CONFLICT,
                )
            self.perform_destroy(count)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def update_counts(self, request, pk=None):
        """Mettre à jour les quantités comptées"""
        payload = InventoryCountUpdateSerializer(data=request.data)
        payload.is_valid(raise_exception=True)

        with transaction.atomic():
            count = get_object_or_404(
                InventoryCount.objects.select_for_update(),
                pk=pk,
            )
            self.check_object_permissions(request, count)
            if count.status != InventoryCount.CountStatus.IN_PROGRESS:
                return Response(
                    {'detail': 'Comptage non en cours.'},
                    status=status.HTTP_409_CONFLICT,
                )

            submitted = payload.validated_data['items']
            item_ids = [entry['id'] for entry in submitted]
            items = {
                item.pk: item
                for item in InventoryCountItem.objects.select_for_update()
                .filter(count=count, pk__in=item_ids)
                .order_by('pk')
            }
            unknown = sorted(set(item_ids) - set(items))
            if unknown:
                return Response(
                    {'items': f'Lignes inconnues pour cet inventaire: {unknown}.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            for entry in submitted:
                item = items[entry['id']]
                item.counted_quantity = entry['counted_quantity']
                item.save(update_fields=['counted_quantity'])

        return Response(InventoryCountSerializer(count, context={'request': request}).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Marquer le comptage comme terminé"""
        with transaction.atomic():
            count = get_object_or_404(
                InventoryCount.objects.select_for_update(),
                pk=pk,
            )
            self.check_object_permissions(request, count)
            if count.status != InventoryCount.CountStatus.IN_PROGRESS:
                return Response(
                    {'detail': 'Seul un comptage en cours peut être terminé.'},
                    status=status.HTTP_409_CONFLICT,
                )
            missing = list(
                count.items.filter(counted_quantity__isnull=True)
                .values_list('id', flat=True)
            )
            if missing:
                return Response(
                    {'items': f'Quantités comptées manquantes pour les lignes: {missing}.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            count.status = InventoryCount.CountStatus.COMPLETED
            count.completed_at = timezone.now()
            count.save(update_fields=['status', 'completed_at'])
        return Response({'status': 'Comptage terminé'})

    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        """Valider le comptage et ajuster le stock"""
        count = get_object_or_404(InventoryCount, pk=pk)
        count, adjustments = validate_inventory_count(count, request.user)

        return Response({
            'status': 'Stock ajusté',
            'adjustments': adjustments
        })
