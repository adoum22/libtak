import os
import subprocess

from django.conf import settings
from django.db import transaction
from drf_spectacular.utils import (
    OpenApiResponse,
    OpenApiTypes,
    extend_schema,
    inline_serializer,
)
from rest_framework import generics, serializers as drf_serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import get_user_model
from django.http import FileResponse, Http404
from django.urls import reverse

from .serializers import (
    UserSerializer,
    MeSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    ChangePasswordSerializer,
    ResetPasswordSerializer,
    AppSettingsSerializer,
    AuditLogSerializer,
    CustomTokenObtainPairSerializer
)
from .models import AppSettings, AuditLog
from .permissions import IsAdminRole, CanManageUsers
from .security import revoke_user_refresh_tokens
from .throttles import FileUploadRateThrottle, LoginAccountRateThrottle
from rest_framework_simplejwt.views import TokenObtainPairView

User = get_user_model()


def usable_active_admins(*, lock=False):
    """Return administrators that can still authenticate to recover the app."""
    queryset = User.objects.filter(
        role=User.Role.ADMIN,
        is_active=True,
    ).exclude(password='').exclude(password__startswith='!')
    return queryset.select_for_update() if lock else queryset


def excel_safe(value):
    """Prevent spreadsheet applications from evaluating untrusted strings."""
    if isinstance(value, str) and value.lstrip().startswith(('=', '+', '-', '@')):
        return "'" + value
    return value


def get_deploy_commit():
    for key in ('VERCEL_GIT_COMMIT_SHA', 'COMMIT_SHA', 'SOURCE_VERSION', 'GIT_COMMIT'):
        value = os.environ.get(key)
        if value:
            return value

    try:
        result = subprocess.run(
            ['git', 'rev-parse', '--short=12', 'HEAD'],
            cwd=settings.BASE_DIR.parent,
            check=True,
            capture_output=True,
            text=True,
            timeout=2,
        )
        return result.stdout.strip()
    except Exception:
        return None


class AppVersionView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses=inline_serializer(
            name='AppVersionResponse',
            fields={
                'backend_commit': drf_serializers.CharField(allow_null=True),
                'backend_commit_short': drf_serializers.CharField(allow_null=True),
                'debug': drf_serializers.BooleanField(),
            },
        )
    )
    def get(self, request):
        commit = get_deploy_commit()
        return Response({
            'backend_commit': commit,
            'backend_commit_short': commit[:12] if commit else None,
            'debug': settings.DEBUG,
        })


class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    throttle_scope = 'login'

    def get_throttles(self):
        if settings.TESTING:
            return super().get_throttles()
        return super().get_throttles() + [LoginAccountRateThrottle()]


class LogoutView(APIView):
    """Blacklist the supplied refresh token."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        request=inline_serializer(
            name='LogoutRequest',
            fields={'refresh': drf_serializers.CharField(required=False)},
        ),
        responses={
            205: OpenApiResponse(description='Session client réinitialisée.'),
        },
    )
    def post(self, request):
        refresh = request.data.get('refresh')
        if refresh:
            try:
                token = RefreshToken(refresh)
                if str(token.get('user_id')) != str(request.user.pk):
                    return Response(
                        {'detail': "Ce jeton n'appartient pas à l'utilisateur connecté."},
                        status=status.HTTP_403_FORBIDDEN,
                    )
                token.blacklist()
            except TokenError:
                return Response(
                    {'detail': 'Jeton de rafraîchissement invalide.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        try:
            AuditLog.log(
                user=request.user,
                action=AuditLog.ActionType.LOGOUT,
                model_name='User',
                object_id=request.user.id,
                object_repr=str(request.user),
                request=request,
            )
        except Exception:
            pass
        return Response(status=status.HTTP_205_RESET_CONTENT)


class UserMeView(generics.RetrieveUpdateAPIView):
    """Vue pour l'utilisateur connecté"""
    serializer_class = MeSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_object(self):
        return self.request.user

    def get_throttles(self):
        throttles = super().get_throttles()
        if not settings.TESTING and self.request.FILES:
            throttles.append(FileUploadRateThrottle())
        return throttles

class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        request=ChangePasswordSerializer,
        responses=inline_serializer(
            name='ChangePasswordResponse',
            fields={'message': drf_serializers.CharField()},
        ),
    )
    def post(self, request):
        """Changer le mot de passe de l'utilisateur connecté"""
        serializer = ChangePasswordSerializer(
            data=request.data,
            context={'request': request},
        )
        serializer.is_valid(raise_exception=True)
        user = request.user
        if not user.check_password(serializer.validated_data['old_password']):
            return Response(
                {'old_password': 'Mot de passe incorrect.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        with transaction.atomic():
            user.set_password(serializer.validated_data['new_password'])
            user.save(update_fields=['password'])
            revoke_user_refresh_tokens(user)
            AuditLog.log(
                user=user,
                action=AuditLog.ActionType.UPDATE,
                model_name='User',
                object_id=user.pk,
                object_repr=str(user),
                changes={'password_changed': True},
                request=request,
            )
        return Response({'message': 'Mot de passe modifié avec succès.'})


class UserViewSet(viewsets.ModelViewSet):
    """API pour la gestion des utilisateurs (Admin only)"""
    queryset = User.objects.all()
    permission_classes = [IsAuthenticated, CanManageUsers]
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    # pagination_class configured in settings.py REST_FRAMEWORK

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        if self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        return UserSerializer

    def get_throttles(self):
        throttles = super().get_throttles()
        if not settings.TESTING and self.request.FILES:
            throttles.append(FileUploadRateThrottle())
        return throttles

    def get_queryset(self):
        queryset = super().get_queryset()
        role = self.request.query_params.get('role')
        if role:
            queryset = queryset.filter(role=role)
        return queryset.order_by('username')

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.log(
            user=self.request.user, action=AuditLog.ActionType.CREATE,
            model_name='User', object_id=instance.id,
            object_repr=str(instance), request=self.request,
        )

    @transaction.atomic
    def perform_update(self, serializer):
        instance = User.objects.select_for_update().get(pk=serializer.instance.pk)
        serializer.instance = instance
        new_role = serializer.validated_data.get('role', instance.role)
        new_is_active = serializer.validated_data.get('is_active', instance.is_active)

        if instance.pk == self.request.user.pk and new_role != User.Role.ADMIN:
            raise ValidationError({'role': 'Vous ne pouvez pas rétrograder votre propre compte.'})
        if instance.pk == self.request.user.pk and not new_is_active:
            raise ValidationError({'is_active': 'Vous ne pouvez pas désactiver votre propre compte.'})

        removes_active_admin = (
            instance.role == User.Role.ADMIN
            and instance.is_active
            and (new_role != User.Role.ADMIN or not new_is_active)
        )
        if removes_active_admin and not usable_active_admins(lock=True).exclude(
            pk=instance.pk
        ).exists():
            raise ValidationError('Le dernier administrateur actif doit être conservé.')

        before = {'role': instance.role, 'is_active': instance.is_active}
        instance = serializer.save()
        if before['is_active'] and not instance.is_active:
            revoke_user_refresh_tokens(instance)
        AuditLog.log(
            user=self.request.user, action=AuditLog.ActionType.UPDATE,
            model_name='User', object_id=instance.id,
            object_repr=str(instance), request=self.request,
            changes={
                'before': before,
                'after': {'role': instance.role, 'is_active': instance.is_active},
            },
        )

    @transaction.atomic
    def perform_destroy(self, instance):
        instance = User.objects.select_for_update().get(pk=instance.pk)
        if instance.pk == self.request.user.pk:
            raise ValidationError('Vous ne pouvez pas supprimer votre propre compte.')
        if (
            instance.role == User.Role.ADMIN
            and instance.is_active
            and not usable_active_admins(lock=True).exclude(pk=instance.pk).exists()
        ):
            raise ValidationError('Le dernier administrateur actif doit être conservé.')
        repr_ = str(instance)
        obj_id = instance.id
        avatar_storage = instance.avatar.storage if instance.avatar else None
        avatar_name = instance.avatar.name if instance.avatar else ''
        revoke_user_refresh_tokens(instance)
        instance.delete()
        if avatar_storage and avatar_name:
            transaction.on_commit(
                lambda: avatar_storage.delete(avatar_name),
                robust=True,
            )
        AuditLog.log(
            user=self.request.user, action=AuditLog.ActionType.DELETE,
            model_name='User', object_id=obj_id,
            object_repr=repr_, request=self.request,
        )

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Réinitialiser le mot de passe d'un utilisateur (Admin)"""
        user = self.get_object()
        serializer = ResetPasswordSerializer(
            data=request.data,
            context={'user': user},
        )
        serializer.is_valid(raise_exception=True)
        with transaction.atomic():
            user.set_password(serializer.validated_data['new_password'])
            user.save(update_fields=['password'])
            revoke_user_refresh_tokens(user)
            AuditLog.log(
                user=request.user,
                action=AuditLog.ActionType.UPDATE,
                model_name='User',
                object_id=user.pk,
                object_repr=str(user),
                changes={'password_reset': True},
                request=request,
            )
        return Response({
            'message': f'Mot de passe réinitialisé pour {user.username}'
        })


    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Activer/Desactiver un utilisateur"""
        with transaction.atomic():
            user = User.objects.select_for_update().get(pk=self.get_object().pk)
            if user.pk == request.user.pk and user.is_active:
                raise ValidationError('Vous ne pouvez pas désactiver votre propre compte.')
            if (
                user.role == User.Role.ADMIN
                and user.is_active
                and not usable_active_admins(lock=True).exclude(pk=user.pk).exists()
            ):
                raise ValidationError('Le dernier administrateur actif doit être conservé.')
            user.is_active = not user.is_active
            user.save(update_fields=['is_active'])
            if not user.is_active:
                revoke_user_refresh_tokens(user)
            AuditLog.log(
                user=request.user,
                action=AuditLog.ActionType.UPDATE,
                model_name='User',
                object_id=user.pk,
                object_repr=str(user),
                changes={'is_active': user.is_active},
                request=request,
            )
        return Response({
            'message': f'Utilisateur {"active" if user.is_active else "desactive"}',
            'is_active': user.is_active
        })


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """Journal d'activite admin en lecture seule."""
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get_queryset(self):
        queryset = AuditLog.objects.select_related('user').all()
        action_name = self.request.query_params.get('action')
        model_name = self.request.query_params.get('model')
        if action_name:
            queryset = queryset.filter(action=action_name)
        if model_name:
            queryset = queryset.filter(model_name__icontains=model_name)
        return queryset

    # Kept private to avoid exposing a meaningless action on audit log rows.
    def _legacy_toggle_active(self, request, pk=None):
        """Activer/Désactiver un utilisateur"""
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        return Response({
            'message': f'Utilisateur {"activé" if user.is_active else "désactivé"}',
            'is_active': user.is_active
        })


class AppSettingsView(generics.RetrieveUpdateAPIView):
    """Vue pour les paramètres de l'application"""
    serializer_class = AppSettingsSerializer
    permission_classes = [IsAuthenticated, IsAdminRole]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_object(self):
        return AppSettings.get_settings()

    def get_throttles(self):
        throttles = super().get_throttles()
        if not settings.TESTING and self.request.FILES:
            throttles.append(FileUploadRateThrottle())
        return throttles


class PublicSettingsView(generics.RetrieveAPIView):
    """Vue publique des paramètres (nom boutique, logo, devise)"""
    serializer_class = AppSettingsSerializer
    permission_classes = [IsAuthenticated]

    def get_object(self):
        return AppSettings.get_settings()

    def retrieve(self, request, *args, **kwargs):
        settings = self.get_object()
        # Retourner seulement les infos publiques
        data = {
            'store_name': settings.store_name,
            'store_address': settings.store_address,
            'store_phone': settings.store_phone,
            'store_email': settings.store_email,
            'currency': settings.currency,
            'currency_symbol': settings.currency_symbol,
            'print_header': settings.print_header,
            'print_footer': settings.print_footer,
            'company_name': settings.company_name,
            'company_rc': settings.company_rc,
            'company_ice': settings.company_ice,
            'company_if': settings.company_if,
            'company_patente': settings.company_patente,
            'company_cnss': settings.company_cnss,
            'invoice_prefix': settings.invoice_prefix,
            'invoice_footer': settings.invoice_footer,
            'logo_url': request.build_absolute_uri(reverse('app_settings_logo')) if settings.store_logo else None
        }
        return Response(data)


class AppSettingsLogoView(APIView):
    """Logo public de la boutique pour affichage et impression."""
    permission_classes = [AllowAny]

    @extend_schema(
        responses={
            200: OpenApiResponse(
                response=OpenApiTypes.BINARY,
                description='Logo de la boutique.',
            ),
            404: OpenApiResponse(description='Aucun logo configuré.'),
        }
    )
    def get(self, request):
        settings = AppSettings.get_settings()
        if not settings.store_logo:
            raise Http404
        return FileResponse(settings.store_logo.open('rb'))


from django.http import HttpResponse
from datetime import datetime
from io import BytesIO

class DatabaseExportView(generics.GenericAPIView):
    """Export de la base de données pour backup en Excel"""
    permission_classes = [IsAuthenticated, IsAdminRole]

    @extend_schema(
        responses={
            (200, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
                OpenApiTypes.BINARY,
        }
    )
    def get(self, request):
        from inventory.models import Product, Category, Supplier
        from sales.models import Sale, SaleItem
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Get selection parameters (default to True if not specified)
        include_products = request.query_params.get('products', 'true').lower() == 'true'
        include_categories = request.query_params.get('categories', 'true').lower() == 'true'
        include_suppliers = request.query_params.get('suppliers', 'true').lower() == 'true'
        include_sales = request.query_params.get('sales', 'true').lower() == 'true'
        include_users = request.query_params.get('users', 'true').lower() == 'true'
        include_settings = request.query_params.get('settings', 'true').lower() == 'true'

        if not any((
            include_products,
            include_categories,
            include_suppliers,
            include_sales,
            include_users,
            include_settings,
        )):
            return Response(
                {'detail': 'Sélectionnez au moins une rubrique à exporter.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Créer le workbook Excel
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        def style_header(ws, headers):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                ws.column_dimensions[cell.column_letter].width = max(15, len(header) + 5)

        # Sheet: Produits
        if include_products:
            ws = wb.active
            ws.title = "Produits"
            headers = ['ID', 'Nom', 'Code-barres', 'Catégorie', 'Fournisseur', 'Prix Achat', 'Prix Vente', 'Stock', 'Seuil', 'Actif']
            style_header(ws, headers)

            for row, prod in enumerate(Product.objects.select_related('category', 'supplier'), 2):
                ws.cell(row=row, column=1, value=prod.id)
                ws.cell(row=row, column=2, value=excel_safe(prod.name))
                ws.cell(row=row, column=3, value=excel_safe(prod.barcode))
                ws.cell(row=row, column=4, value=excel_safe(prod.category.name if prod.category else ''))
                ws.cell(row=row, column=5, value=excel_safe(prod.supplier.name if prod.supplier else ''))
                ws.cell(row=row, column=6, value=float(prod.purchase_price))
                ws.cell(row=row, column=7, value=float(prod.sale_price_ht))
                ws.cell(row=row, column=8, value=prod.stock)
                ws.cell(row=row, column=9, value=prod.min_stock)
                ws.cell(row=row, column=10, value='Oui' if prod.active else 'Non')

        # Sheet: Catégories
        if include_categories:
            ws = wb.create_sheet("Catégories")
            headers = ['ID', 'Nom', 'Description']
            style_header(ws, headers)

            for row, cat in enumerate(Category.objects.all(), 2):
                ws.cell(row=row, column=1, value=cat.id)
                ws.cell(row=row, column=2, value=excel_safe(cat.name))
                ws.cell(row=row, column=3, value=excel_safe(getattr(cat, 'description', '')))

        # Sheet: Fournisseurs
        if include_suppliers:
            ws = wb.create_sheet("Fournisseurs")
            headers = ['ID', 'Nom', 'Contact', 'Email', 'Téléphone', 'Adresse', 'Notes']
            style_header(ws, headers)

            for row, sup in enumerate(Supplier.objects.all(), 2):
                ws.cell(row=row, column=1, value=sup.id)
                ws.cell(row=row, column=2, value=excel_safe(sup.name))
                ws.cell(row=row, column=3, value=excel_safe(sup.contact_name))
                ws.cell(row=row, column=4, value=excel_safe(sup.email))
                ws.cell(row=row, column=5, value=excel_safe(sup.phone))
                ws.cell(row=row, column=6, value=excel_safe(sup.address))
                ws.cell(row=row, column=7, value=excel_safe(sup.notes))

        # Sheet: Ventes
        if include_sales:
            ws = wb.create_sheet("Ventes")
            headers = ['ID Vente', 'Date', 'Total', 'Mode Paiement', 'Caissier', 'Produit', 'Quantité', 'Prix Unit.', 'Sous-total']
            style_header(ws, headers)

            row = 2
            sales_qs = (
                Sale.objects.select_related('user')
                .prefetch_related('items__product')
                .order_by('-created_at')[:1000]
            )
            for sale in sales_qs:
                for item in sale.items.all():
                    ws.cell(row=row, column=1, value=sale.id)
                    ws.cell(row=row, column=2, value=sale.created_at.strftime('%Y-%m-%d %H:%M'))
                    ws.cell(row=row, column=3, value=float(sale.total_ttc))
                    ws.cell(row=row, column=4, value=excel_safe(sale.payment_method))
                    ws.cell(row=row, column=5, value=excel_safe(sale.user.username if sale.user else ''))
                    ws.cell(row=row, column=6, value=excel_safe(item.product.name if item.product else (item.product_name or 'Produit supprimé')))
                    ws.cell(row=row, column=7, value=item.quantity)
                    ws.cell(row=row, column=8, value=float(item.unit_price_ht))
                    ws.cell(row=row, column=9, value=float(item.total_price_ht))
                    row += 1

        # Sheet: Utilisateurs
        if include_users:
            ws = wb.create_sheet("Utilisateurs")
            headers = ['ID', 'Nom utilisateur', 'Email', 'Prénom', 'Nom', 'Rôle', 'Téléphone', 'Actif', 'Voir Stock', 'Gérer Stock']
            style_header(ws, headers)

            for row, user in enumerate(User.objects.all(), 2):
                ws.cell(row=row, column=1, value=user.id)
                ws.cell(row=row, column=2, value=excel_safe(user.username))
                ws.cell(row=row, column=3, value=excel_safe(user.email))
                ws.cell(row=row, column=4, value=excel_safe(user.first_name))
                ws.cell(row=row, column=5, value=excel_safe(user.last_name))
                ws.cell(row=row, column=6, value=excel_safe(user.role))
                ws.cell(row=row, column=7, value=excel_safe(user.phone))
                ws.cell(row=row, column=8, value='Oui' if user.is_active else 'Non')
                ws.cell(row=row, column=9, value='Oui' if user.can_view_stock else 'Non')
                ws.cell(row=row, column=10, value='Oui' if user.can_manage_stock else 'Non')

        # Sheet: Paramètres
        if include_settings:
            ws = wb.create_sheet("Paramètres")
            settings = AppSettings.get_settings()

            ws.cell(row=1, column=1, value="Paramètre").font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=2, value="Valeur").font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 40

            params = [
                ('Nom de la boutique', settings.store_name),
                ('Adresse', settings.store_address),
                ('Téléphone', settings.store_phone),
                ('Email', settings.store_email),
                ('Devise', settings.currency),
                ('Symbole devise', settings.currency_symbol),
                ('Date export', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ]

            for row, (param, value) in enumerate(params, 2):
                ws.cell(row=row, column=1, value=param)
                ws.cell(row=row, column=2, value=excel_safe(value))

        # Si pas de produits sélectionnés, supprimer la feuille par défaut vide
        if not include_products and 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Créer la réponse HTTP avec le fichier Excel
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        AuditLog.log(
            user=request.user,
            action=AuditLog.ActionType.EXPORT,
            model_name='DatabaseExport',
            object_repr='Excel data export',
            changes={
                'products': include_products,
                'categories': include_categories,
                'suppliers': include_suppliers,
                'sales': include_sales,
                'users': include_users,
                'settings': include_settings,
            },
            request=request,
        )

        filename = f"libtak_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
