import base64
import os
import sqlite3
import ssl
import tempfile
import zipfile
from contextlib import closing
from datetime import timedelta
from decimal import Decimal
from pathlib import Path

from django.test import (
    RequestFactory,
    SimpleTestCase,
    TestCase,
    TransactionTestCase,
    override_settings,
)
from django.contrib.auth import get_user_model
from django.contrib.auth.models import AnonymousUser
from django.conf import settings
from django.utils import timezone
from unittest.mock import MagicMock, patch
from asgiref.sync import async_to_sync
from channels.testing import WebsocketCommunicator
from rest_framework.test import APITestCase
from rest_framework import status
from rest_framework.settings import api_settings

from .consumers import StockConsumer
from .models import AppSettings, AuditLog
from .sync_api import _import_return, _import_sale
from .sync_service import SyncService, make_sync_id
from .serializers import ChangePasswordSerializer, UserCreateSerializer
from .views import excel_safe
from create_users import bootstrap_admin, initialize_app_settings
from send_reports import backup_database, send_email
from reporting.backup_utils import decrypt_archive, validate_zip_archive
from inventory.models import Product, StockMovement
from sales.models import Return, ReturnItem, Sale, SaleItem

User = get_user_model()


class UserModelTest(TestCase):
    """Tests pour le modèle User personnalisé"""

    def test_create_admin_user(self):
        """Test création utilisateur admin"""
        user = User.objects.create_user(
            username='testadmin',
            email='admin@test.com',
            password='testpass123',
            role='ADMIN'
        )
        self.assertEqual(user.username, 'testadmin')
        self.assertEqual(user.role, 'ADMIN')
        self.assertTrue(user.is_active)
        self.assertTrue(user.is_staff)

    def test_create_cashier_user(self):
        """Test création utilisateur caissier"""
        user = User.objects.create_user(
            username='testcashier',
            email='cashier@test.com',
            password='testpass123',
            role='CASHIER'
        )
        self.assertEqual(user.role, 'CASHIER')
        self.assertFalse(user.can_manage_stock)
        self.assertFalse(user.is_staff)

    def test_demoting_user_removes_django_admin_privileges(self):
        user = User.objects.create_superuser(
            username='former-admin',
            email='former@example.com',
            password='A-long-random-passphrase-2026',
        )
        user.role = User.Role.CASHIER
        user.save(update_fields=['role'])
        user.refresh_from_db()
        self.assertFalse(user.is_staff)
        self.assertFalse(user.is_superuser)

    def test_user_permissions(self):
        """Test permissions utilisateur"""
        user = User.objects.create_user(
            username='testperm',
            password='testpass123',
            can_view_stock=True,
            can_manage_stock=False
        )
        self.assertTrue(user.can_view_stock)
        self.assertFalse(user.can_manage_stock)


class AuthenticationAPITest(APITestCase):
    """Tests pour l'API d'authentification"""

    def setUp(self):
        self.user = User.objects.create_user(
            username='testuser',
            email='test@test.com',
            password='testpass123',
            role='ADMIN'
        )

    def test_login_success(self):
        """Test connexion réussie"""
        response = self.client.post('/api/auth/login/', {
            'username': 'testuser',
            'password': 'testpass123'
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('access', response.data)
        self.assertIn('refresh', response.data)

    def test_login_invalid_credentials(self):
        """Test connexion avec mauvais mot de passe"""
        response = self.client.post('/api/auth/login/', {
            'username': 'testuser',
            'password': 'wrongpassword'
        })
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_login_by_email_is_case_insensitive(self):
        response = self.client.post('/api/auth/login/', {
            'username': 'TEST@TEST.COM',
            'password': 'testpass123',
        })
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('access', response.data)

    def test_protected_endpoint_without_token(self):
        """Test accès endpoint protégé sans token"""
        response = self.client.get('/api/auth/me/')
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_protected_endpoint_with_token(self):
        """Test accès endpoint protégé avec token"""
        # Get token
        login_response = self.client.post('/api/auth/login/', {
            'username': 'testuser',
            'password': 'testpass123'
        })
        token = login_response.data['access']

        # Access protected endpoint
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        response = self.client.get('/api/auth/me/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['username'], 'testuser')

    def test_me_returns_effective_global_cashier_stock_permissions(self):
        cashier = User.objects.create_user(
            username='cashier-global-stock',
            password='cashier123',
            role='CASHIER',
            can_view_stock=False,
            can_manage_stock=False,
        )
        app_settings = AppSettings.get_settings()
        app_settings.cashier_can_manage_stock = True
        app_settings.cashier_can_view_stock = False
        app_settings.save(update_fields=[
            'cashier_can_manage_stock',
            'cashier_can_view_stock',
        ])
        self.client.credentials()
        self.client.force_authenticate(user=cashier)

        response = self.client.get('/api/auth/me/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.data['can_manage_stock'])
        self.assertTrue(response.data['can_view_stock'])

    def test_me_patch_cannot_escalate_role_or_permissions(self):
        """Un utilisateur ne peut pas modifier ses propres droits via /me/."""
        cashier = User.objects.create_user(
            username='cashier',
            password='cashier123',
            role='CASHIER',
            can_view_stock=False,
            can_manage_stock=False,
        )
        self.client.credentials()
        self.client.force_authenticate(user=cashier)

        response = self.client.patch('/api/auth/me/', {
            'role': 'ADMIN',
            'can_view_stock': True,
            'can_manage_stock': True,
            'is_active': False,
            'first_name': 'Updated',
        }, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        cashier.refresh_from_db()
        self.assertEqual(cashier.role, 'CASHIER')
        self.assertFalse(cashier.can_view_stock)
        self.assertFalse(cashier.can_manage_stock)
        self.assertTrue(cashier.is_active)
        self.assertEqual(cashier.first_name, 'Updated')

    def test_change_password_route(self):
        self.client.credentials()
        self.client.force_authenticate(user=self.user)
        response = self.client.post('/api/auth/me/change-password/', {
            'old_password': 'testpass123',
            'new_password': 'newpass123-strong',
            'new_password_confirm': 'newpass123-strong',
        }, format='json')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.user.refresh_from_db()
        self.assertTrue(self.user.check_password('newpass123-strong'))

    def test_change_password_revokes_existing_access_and_refresh_tokens(self):
        login = self.client.post('/api/auth/login/', {
            'username': 'testuser',
            'password': 'testpass123',
        })
        access = login.data['access']
        refresh = login.data['refresh']
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {access}')

        changed = self.client.post('/api/auth/me/change-password/', {
            'old_password': 'testpass123',
            'new_password': 'A-new-unique-passphrase-2026!',
            'new_password_confirm': 'A-new-unique-passphrase-2026!',
        }, format='json')
        self.assertEqual(changed.status_code, status.HTTP_200_OK)

        self.assertEqual(
            self.client.get('/api/auth/me/').status_code,
            status.HTTP_401_UNAUTHORIZED,
        )
        self.client.credentials()
        self.assertEqual(
            self.client.post('/api/auth/refresh/', {'refresh': refresh}).status_code,
            status.HTTP_401_UNAUTHORIZED,
        )

    def test_common_password_is_rejected_by_all_password_serializers(self):
        create_serializer = UserCreateSerializer(data={
            'username': 'someone',
            'password': 'password1234',
            'password_confirm': 'password1234',
            'role': 'CASHIER',
        })
        self.assertFalse(create_serializer.is_valid())
        self.assertIn('password', create_serializer.errors)

        request = type('Request', (), {'user': self.user})()
        change_serializer = ChangePasswordSerializer(
            data={
                'old_password': 'testpass123',
                'new_password': 'password1234',
                'new_password_confirm': 'password1234',
            },
            context={'request': request},
        )
        self.assertFalse(change_serializer.is_valid())
        self.assertIn('new_password', change_serializer.errors)


class UserAPITest(APITestCase):
    """Tests pour l'API de gestion des utilisateurs"""

    def setUp(self):
        self.admin = User.objects.create_user(
            username='admin',
            password='admin123',
            role='ADMIN'
        )
        # Authenticate as admin
        login_response = self.client.post('/api/auth/login/', {
            'username': 'admin',
            'password': 'admin123'
        })
        self.token = login_response.data['access']
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.token}')

    def test_list_users(self):
        """Test liste des utilisateurs"""
        response = self.client.get('/api/auth/users/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)

    def test_create_user(self):
        """Test création utilisateur via API"""
        response = self.client.post('/api/auth/users/', {
            'username': 'newuser',
            'email': 'new@test.com',
            'password': 'newpass123-strong',
            'password_confirm': 'newpass123-strong',
            'first_name': 'New',
            'last_name': 'User',
            'role': 'CASHIER'
        })
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(User.objects.count(), 2)

    def test_admin_cannot_demote_or_delete_own_account(self):
        demote = self.client.patch(
            f'/api/auth/users/{self.admin.pk}/',
            {'role': 'CASHIER'},
            format='json',
        )
        self.assertEqual(demote.status_code, status.HTTP_400_BAD_REQUEST)
        self.admin.refresh_from_db()
        self.assertEqual(self.admin.role, User.Role.ADMIN)
        self.assertTrue(self.admin.is_staff)

        delete = self.client.delete(f'/api/auth/users/{self.admin.pk}/')
        self.assertEqual(delete.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(User.objects.filter(pk=self.admin.pk).exists())

    def test_admin_cannot_deactivate_own_account(self):
        response = self.client.post(
            f'/api/auth/users/{self.admin.pk}/toggle_active/',
            {},
            format='json',
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.admin.refresh_from_db()
        self.assertTrue(self.admin.is_active)

    def test_reset_password_enforces_policy_and_revokes_refresh(self):
        cashier = User.objects.create_user(
            username='cashier-reset',
            password='Initial-long-passphrase-2026!',
            role=User.Role.CASHIER,
        )
        login = self.client.post('/api/auth/login/', {
            'username': cashier.username,
            'password': 'Initial-long-passphrase-2026!',
        })
        refresh = login.data['refresh']
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {self.token}')

        weak = self.client.post(
            f'/api/auth/users/{cashier.pk}/reset_password/',
            {'new_password': 'password1234'},
            format='json',
        )
        self.assertEqual(weak.status_code, status.HTTP_400_BAD_REQUEST)

        strong = self.client.post(
            f'/api/auth/users/{cashier.pk}/reset_password/',
            {'new_password': 'Replacement-passphrase-2026!'},
            format='json',
        )
        self.assertEqual(strong.status_code, status.HTTP_200_OK)
        self.client.credentials()
        self.assertEqual(
            self.client.post('/api/auth/refresh/', {'refresh': refresh}).status_code,
            status.HTTP_401_UNAUTHORIZED,
        )


class SecurityDefaultsTest(TestCase):
    def test_drf_is_authenticated_by_default(self):
        permission_names = {
            permission.__name__ for permission in api_settings.DEFAULT_PERMISSION_CLASSES
        }
        self.assertEqual(permission_names, {'IsAuthenticated'})

    def test_jwt_password_change_revocation_is_enabled(self):
        self.assertTrue(settings.SIMPLE_JWT['CHECK_REVOKE_TOKEN'])

    def test_throttling_and_audit_use_same_proxy_boundary(self):
        self.assertEqual(
            api_settings.NUM_PROXIES,
            settings.AUDIT_TRUSTED_PROXY_COUNT,
        )

    @override_settings(AUDIT_TRUSTED_PROXY_COUNT=0)
    def test_audit_does_not_trust_forwarded_ip_by_default(self):
        request = RequestFactory().get(
            '/',
            REMOTE_ADDR='127.0.0.1',
            HTTP_X_FORWARDED_FOR='203.0.113.9',
        )
        event = AuditLog.log(
            user=None,
            action=AuditLog.ActionType.LOGIN,
            model_name='User',
            request=request,
        )
        self.assertEqual(str(event.ip_address), '127.0.0.1')

    def test_api_responses_include_restrictive_browser_headers(self):
        response = self.client.get('/api/health/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("default-src 'none'", response['Content-Security-Policy'])
        self.assertEqual(response['X-Frame-Options'], 'DENY')

        logo_response = self.client.get('/api/auth/settings/logo/')
        self.assertEqual(
            logo_response['Cross-Origin-Resource-Policy'],
            'cross-origin',
        )

    def test_excel_formula_prefixes_are_escaped(self):
        for value in ('=1+1', '+SUM(A1:A2)', '-2+3', '@malicious'):
            self.assertEqual(excel_safe(value), "'" + value)
        self.assertEqual(excel_safe('ordinary text'), 'ordinary text')

    def test_database_export_is_valid_filtered_workbook(self):
        from io import BytesIO
        from openpyxl import load_workbook

        admin = User.objects.create_user(
            username='export-admin',
            password='Safe-Export-Password-2026!',
            role=User.Role.ADMIN,
        )
        Product.objects.create(
            name='=FORMULA()',
            barcode='export-001',
            purchase_price=Decimal('2.00'),
            sale_price_ht=Decimal('3.00'),
            stock=4,
        )
        from rest_framework_simplejwt.tokens import RefreshToken
        authorization = f'Bearer {RefreshToken.for_user(admin).access_token}'

        response = self.client.get(
            '/api/auth/backup/',
            {
                'products': 'true',
                'categories': 'false',
                'suppliers': 'false',
                'sales': 'false',
                'users': 'false',
                'settings': 'false',
            },
            HTTP_AUTHORIZATION=authorization,
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ['Produits'])
        self.assertEqual(workbook['Produits']['B2'].value, "'=FORMULA()")

    def test_database_export_rejects_empty_selection(self):
        admin = User.objects.create_user(
            username='empty-export-admin',
            password='Safe-Export-Password-2026!',
            role=User.Role.ADMIN,
        )
        from rest_framework_simplejwt.tokens import RefreshToken
        authorization = f'Bearer {RefreshToken.for_user(admin).access_token}'
        response = self.client.get(
            '/api/auth/backup/?products=false&categories=false&suppliers=false'
            '&sales=false&users=false&settings=false',
            HTTP_AUTHORIZATION=authorization,
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_database_export_is_admin_only(self):
        cashier = User.objects.create_user(
            username='export-cashier',
            password='Safe-Export-Password-2026!',
            role=User.Role.CASHIER,
        )
        from rest_framework_simplejwt.tokens import RefreshToken
        authorization = f'Bearer {RefreshToken.for_user(cashier).access_token}'
        response = self.client.get(
            '/api/auth/backup/',
            HTTP_AUTHORIZATION=authorization,
        )
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_anonymous_websocket_is_rejected(self):
        async def connect():
            communicator = WebsocketCommunicator(
                StockConsumer.as_asgi(),
                '/ws/stock/',
            )
            communicator.scope['user'] = AnonymousUser()
            connected, close_code = await communicator.connect()
            await communicator.disconnect()
            return connected, close_code

        connected, close_code = async_to_sync(connect)()
        self.assertFalse(connected)
        self.assertEqual(close_code, 4401)


class StockConsumerPermissionTest(TransactionTestCase):
    def _connect(self, user):
        async def connect():
            communicator = WebsocketCommunicator(
                StockConsumer.as_asgi(),
                '/ws/stock/',
            )
            communicator.scope['user'] = user
            connected, close_code = await communicator.connect()
            if connected:
                await communicator.disconnect()
            return connected, close_code

        return async_to_sync(connect)()

    def test_cashier_without_effective_stock_visibility_is_rejected(self):
        cashier = User.objects.create_user(
            username='websocket-no-stock',
            password='Safe-Websocket-Password-2026!',
            role=User.Role.CASHIER,
            can_view_stock=False,
            can_manage_stock=False,
        )

        connected, close_code = self._connect(cashier)

        self.assertFalse(connected)
        self.assertEqual(close_code, 4403)

    def test_cashier_with_individual_stock_visibility_is_accepted(self):
        cashier = User.objects.create_user(
            username='websocket-view-stock',
            password='Safe-Websocket-Password-2026!',
            role=User.Role.CASHIER,
            can_view_stock=True,
        )

        connected, close_code = self._connect(cashier)

        self.assertTrue(connected)
        self.assertIsNone(close_code)

    def test_global_stock_visibility_is_honored_by_websocket(self):
        app_settings = AppSettings.get_settings()
        app_settings.cashier_can_view_stock = True
        app_settings.save(update_fields=['cashier_can_view_stock'])
        cashier = User.objects.create_user(
            username='websocket-global-stock',
            password='Safe-Websocket-Password-2026!',
            role=User.Role.CASHIER,
            can_view_stock=False,
        )

        connected, close_code = self._connect(cashier)

        self.assertTrue(connected)
        self.assertIsNone(close_code)


class ContainerSecurityConfigTest(SimpleTestCase):
    def test_runtime_is_non_root_and_build_context_excludes_secrets(self):
        dockerfile = (settings.BASE_DIR / 'Dockerfile').read_text(encoding='utf-8')
        dockerignore = (settings.BASE_DIR / '.dockerignore').read_text(
            encoding='utf-8'
        ).splitlines()

        self.assertIn('USER 10001:10001', dockerfile)
        self.assertIn('.env', dockerignore)
        self.assertIn('db.sqlite3*', dockerignore)
        self.assertIn('*.sqlite3', dockerignore)
        self.assertIn('venv', dockerignore)
        self.assertIn('.venv', dockerignore)

    def test_compose_uses_read_only_application_containers(self):
        compose = (settings.BASE_DIR.parent / 'docker-compose.yml').read_text(
            encoding='utf-8'
        )
        self.assertIn('read_only: true', compose)
        self.assertIn('user: "10001:10001"', compose)
        self.assertIn('127.0.0.1:8000:8000', compose)


class BootstrapSecurityTest(TestCase):
    def test_bootstrap_requires_operator_supplied_credentials(self):
        with patch.dict(os.environ, {
            'BOOTSTRAP_ADMIN_USERNAME': '',
            'BOOTSTRAP_ADMIN_PASSWORD': '',
            'BOOTSTRAP_ADMIN_EMAIL': '',
        }):
            with self.assertRaises(RuntimeError):
                bootstrap_admin()

    def test_bootstrap_is_idempotent_and_does_not_overwrite_settings(self):
        environment = {
            'BOOTSTRAP_ADMIN_USERNAME': 'bootstrap-owner',
            'BOOTSTRAP_ADMIN_PASSWORD': 'Unique-bootstrap-passphrase-2026!',
            'BOOTSTRAP_ADMIN_EMAIL': 'owner@example.com',
        }
        with patch.dict(os.environ, environment):
            user, created = bootstrap_admin()
            self.assertTrue(created)
            original_hash = user.password

            user_again, created_again = bootstrap_admin()
            self.assertFalse(created_again)
            self.assertEqual(user_again.pk, user.pk)
            self.assertEqual(user_again.password, original_hash)

        app_settings = AppSettings.objects.create(
            pk=1,
            store_name='Configuration personnalisée',
        )
        initialize_app_settings()
        app_settings.refresh_from_db()
        self.assertEqual(app_settings.store_name, 'Configuration personnalisée')


class ReportTransportSecurityTest(SimpleTestCase):
    @override_settings(
        EMAIL_HOST='smtp.example.com',
        EMAIL_PORT=587,
        EMAIL_HOST_USER='mailer@example.com',
        EMAIL_HOST_PASSWORD='not-a-real-secret',
        DEFAULT_FROM_EMAIL='mailer@example.com',
        EMAIL_USE_TLS=True,
        EMAIL_USE_SSL=False,
        EMAIL_TIMEOUT=15,
    )
    def test_starttls_uses_a_verifying_ssl_context(self):
        report_settings = MagicMock()
        report_settings.get_recipients_list.return_value = ['owner@example.com']
        server = MagicMock()

        environment = {
            'EMAIL_HOST': 'smtp.example.com',
            'EMAIL_PORT': '587',
            'EMAIL_HOST_USER': 'mailer@example.com',
            'EMAIL_HOST_PASSWORD': 'not-a-real-secret',
            'DEFAULT_FROM_EMAIL': 'mailer@example.com',
        }
        with patch.dict(os.environ, environment), patch(
            'send_reports.smtplib.SMTP'
        ) as smtp:
            smtp.return_value.__enter__.return_value = server
            self.assertTrue(send_email(report_settings, 'subject', '<p>body</p>'))

        context = server.starttls.call_args.kwargs['context']
        self.assertTrue(context.check_hostname)
        self.assertEqual(context.verify_mode, ssl.CERT_REQUIRED)
        smtp.assert_called_once_with('smtp.example.com', 587, timeout=15)

    def test_sqlite_backup_is_local_and_consistent(self):
        with tempfile.TemporaryDirectory() as directory:
            directory_path = Path(directory)
            source_path = directory_path / 'source.sqlite3'
            backup_dir = directory_path / 'private-backups'
            with closing(sqlite3.connect(source_path)) as connection:
                connection.execute('CREATE TABLE sample (value TEXT)')
                connection.execute('INSERT INTO sample VALUES (?)', ('preserved',))
                connection.commit()

            database_settings = {
                'default': {
                    'ENGINE': 'django.db.backends.sqlite3',
                    'NAME': source_path,
                }
            }
            encrypted_key = base64.urlsafe_b64encode(b'k' * 32).decode('ascii')
            with override_settings(
                DATABASES=database_settings,
                MEDIA_ROOT=directory_path / 'empty-media',
            ), patch.dict(os.environ, {
                'LIBTAK_BACKUP_DIR': str(backup_dir),
                'BACKUP_ENCRYPTION_KEY': encrypted_key,
                'BACKUP_RETENTION_DAYS': '2',
            }), patch('reporting.tasks.ReportLog.objects.create'):
                backup_path = backup_database()

            self.assertIsNotNone(backup_path)
            self.assertTrue(Path(backup_path).is_file())
            self.assertEqual(Path(backup_path).suffix, '.ltbk')
            decrypted = directory_path / 'decrypted.zip'
            with patch.dict(
                os.environ,
                {'BACKUP_ENCRYPTION_KEY': encrypted_key},
            ):
                decrypt_archive(backup_path, decrypted)
            manifest = validate_zip_archive(decrypted)
            self.assertEqual(manifest['database_engine'], 'django.db.backends.sqlite3')
            restored = directory_path / 'restored.sqlite3'
            with zipfile.ZipFile(decrypted, 'r') as archive:
                restored.write_bytes(archive.read('database.sqlite3'))
            with closing(sqlite3.connect(restored)) as backup_connection:
                value = backup_connection.execute(
                    'SELECT value FROM sample'
                ).fetchone()[0]
            self.assertEqual(value, 'preserved')


class SyncRoundTripRegressionTest(TestCase):
    origin_id = '123e4567-e89b-12d3-a456-426614174000'

    def setUp(self):
        self.user = User.objects.create_user(
            username='sync-operator',
            password='A-long-sync-test-passphrase-2026!',
            role=User.Role.ADMIN,
        )
        self.product = Product.objects.create(
            name='Sync product',
            barcode='SYNC-ROUNDTRIP-1',
            purchase_price=Decimal('4.00'),
            sale_price_ht=Decimal('10.00'),
            tva=Decimal('20.00'),
            stock=17,
        )

    def _create_source_sale(self, payload_hash='a' * 64):
        sale = Sale.objects.create(
            user=self.user,
            total_ht=Decimal('10.00'),
            total_tva=Decimal('2.00'),
            total_ttc=Decimal('12.00'),
            discount_amount=Decimal('0.00'),
            payment_method=Sale.PaymentMethod.CASH,
            amount_received=Decimal('20.00'),
            change_amount=Decimal('8.00'),
            idempotency_payload_hash=payload_hash,
        )
        item = SaleItem.objects.create(
            sale=sale,
            product=self.product,
            product_name=self.product.name,
            quantity=1,
            unit_price_ht=Decimal('10.00'),
            total_price_ht=Decimal('10.00'),
            tva_rate=Decimal('20.00'),
            unit_purchase_price=Decimal('4.00'),
            total_purchase_cost=Decimal('4.00'),
        )
        return sale, item

    def test_sale_and_return_idempotency_hashes_survive_round_trip(self):
        source_sale, source_item = self._create_source_sale()
        source_return = Return.objects.create(
            sale=source_sale,
            status=Return.ReturnStatus.PENDING,
            reason='Round-trip test',
            refund_amount=Decimal('12.00'),
            refund_method=Sale.PaymentMethod.CASH,
            processed_by=self.user,
            idempotency_payload_hash='b' * 64,
        )
        ReturnItem.objects.create(
            return_order=source_return,
            sale_item=source_item,
            quantity=1,
            restock=True,
        )
        service = SyncService(origin_id=self.origin_id)

        sale_payload = service._serialize_sale(source_sale, self.origin_id)
        sale_action, _ = _import_sale(sale_payload, self.origin_id)
        self.assertEqual(sale_action, 'created')
        imported_sale = Sale.objects.get(local_sync_id=sale_payload['sync_id'])
        self.assertEqual(
            imported_sale.idempotency_payload_hash,
            source_sale.idempotency_payload_hash,
        )

        return_payload = service._serialize_return(source_return, self.origin_id)
        return_action, _ = _import_return(return_payload, self.origin_id)
        self.assertEqual(return_action, 'created')
        imported_return = Return.objects.get(local_sync_id=return_payload['sync_id'])
        self.assertEqual(
            imported_return.idempotency_payload_hash,
            source_return.idempotency_payload_hash,
        )

    def test_imported_return_statuses_never_apply_inventory_side_effects(self):
        source_sale, source_item = self._create_source_sale()
        source_sale.local_sync_id = make_sync_id(self.origin_id, 'sale', 101)
        source_sale.save(update_fields=['local_sync_id'])
        initial_stock = self.product.stock
        initial_movement_count = StockMovement.objects.count()
        timestamp = timezone.now()
        payload = {
            'local_id': '202',
            'sync_id': make_sync_id(self.origin_id, 'return', 202),
            'sale_local_id': '101',
            'sale_sync_id': source_sale.local_sync_id,
            'reason': 'Status replication only',
            'refund_amount': '12.00',
            'refund_method': Sale.PaymentMethod.CASH,
            'created_at': timestamp.isoformat(),
            'updated_at': timestamp.isoformat(),
            'stock_restored_at': None,
            'completed_at': None,
            'processed_by_username': self.user.username,
            'items': [{
                'sale_item_index': 0,
                'sale_item_local_id': str(source_item.pk),
                'product_barcode': self.product.barcode,
                'product_name': self.product.name,
                'quantity': 1,
                'restock': True,
            }],
        }

        expected_actions = ('created', 'updated', 'updated')
        for index, (return_status, expected_action) in enumerate(zip(
            (
                Return.ReturnStatus.PENDING,
                Return.ReturnStatus.APPROVED,
                Return.ReturnStatus.COMPLETED,
            ),
            expected_actions,
        )):
            payload['status'] = return_status
            payload['idempotency_payload_hash'] = str(index + 1) * 64
            payload['updated_at'] = (
                timestamp + timedelta(seconds=index)
            ).isoformat()
            if return_status == Return.ReturnStatus.COMPLETED:
                payload['completed_at'] = payload['updated_at']

            action, _ = _import_return(dict(payload), self.origin_id)
            self.assertEqual(action, expected_action)
            self.product.refresh_from_db()
            self.assertEqual(self.product.stock, initial_stock)
            self.assertEqual(StockMovement.objects.count(), initial_movement_count)
